import streamlit as st
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import io
import re
import codecs
from datetime import datetime
from collections import Counter

# ── Page config ───────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Data Processing Toolbox",
    page_icon="🚛",
    layout="wide"
)

# ── Theme: bright, modern, friendly ───────────────────────────────────────────
friendly_theme = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Nunito:wght@400;600;700;800;900&display=swap');

#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}

/* ---- Base ---- */
.stApp {
    background: linear-gradient(160deg, #eef4ff 0%, #f4eeff 42%, #fdeff4 72%, #fff3ec 100%);
    background-attachment: fixed;
    color: #3b4256;
}
html, body, [class*="css"], button, input, textarea, select {
    font-family: 'Nunito', 'Segoe UI', system-ui, -apple-system, sans-serif;
}
.block-container { padding-top: 1.4rem; padding-left: 2.2rem; padding-right: 2.2rem; }
h1, h2, h3 { color: #2b2350; font-weight: 800; letter-spacing: -0.4px; }
[data-testid="stMarkdownContainer"] p, label, li { color: #4a5169; }

/* ---- Hero banner ---- */
.app-hero {
    display: flex; align-items: center; gap: 18px;
    background: #ffffff; border: 1px solid #efe9ff; border-radius: 22px;
    padding: 20px 26px; margin-bottom: 20px;
    box-shadow: 0 10px 30px rgba(124,108,246,.12);
}
.app-hero .emoji {
    font-size: 30px; line-height: 1; flex: 0 0 58px; width: 58px; height: 58px;
    display: flex; align-items: center; justify-content: center; border-radius: 18px;
    background: linear-gradient(135deg, #7c6cf6 0%, #b06bf0 55%, #ff85b3 100%);
    box-shadow: 0 6px 16px rgba(160,90,230,.35);
}
.app-hero .title {
    font-size: 26px; font-weight: 900; letter-spacing: -0.6px; line-height: 1.1;
    background: linear-gradient(135deg, #6d5cf0 0%, #b06bf0 55%, #ff6fa5 100%);
    -webkit-background-clip: text; background-clip: text; -webkit-text-fill-color: transparent;
}
.app-hero .sub { color: #6b7189; font-size: 14.5px; font-weight: 700; margin-top: 3px; }

/* ---- Tabs ---- */
.stTabs [data-baseweb="tab-list"] { gap: 8px; background: transparent; border-bottom: none; flex-wrap: wrap; }
.stTabs [data-baseweb="tab"] {
    background: #ffffffcc; border-radius: 14px; padding: 9px 16px;
    border: 1px solid #ece7fb; font-weight: 700; color: #6a6f86;
    box-shadow: 0 1px 2px rgba(40,30,90,.05); transition: all .12s ease;
}
.stTabs [data-baseweb="tab"]:hover { transform: translateY(-1px); color: #5b4fe0; border-color: #d9d0fb; }
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, #7c6cf6 0%, #a86bef 100%);
    color: #ffffff !important; border: 1px solid transparent;
    box-shadow: 0 6px 16px rgba(124,108,246,.38);
}
.stTabs [data-baseweb="tab-highlight"], .stTabs [data-baseweb="tab-border"] { background: transparent; }

/* ---- Buttons ---- */
.stDownloadButton button, .stButton button {
    background: linear-gradient(135deg, #7c6cf6 0%, #b06bf0 100%);
    color: #ffffff; border: none; border-radius: 13px;
    padding: .58rem 1.2rem; font-weight: 800; letter-spacing: .2px;
    box-shadow: 0 6px 16px rgba(124,108,246,.32);
    transition: filter .12s ease, transform .06s ease, box-shadow .12s ease;
}
.stDownloadButton button:hover, .stButton button:hover {
    filter: brightness(1.05); transform: translateY(-1px);
    box-shadow: 0 10px 22px rgba(124,108,246,.42);
}

/* ---- Metric cards ---- */
[data-testid="stMetric"] {
    background: #ffffff; border: 1px solid #efe9ff; border-radius: 18px;
    padding: 16px 20px; box-shadow: 0 6px 18px rgba(124,108,246,.10);
}
[data-testid="stMetricValue"] { color: #2b2350; font-weight: 900; }
[data-testid="stMetricLabel"] { color: #8a8fa6; font-weight: 700; }

/* ---- File uploader ---- */
[data-testid="stFileUploader"] { background: transparent; }
[data-testid="stFileUploader"] section {
    border: 2px dashed #cbb8fb; border-radius: 18px;
    background: linear-gradient(180deg, #fbf9ff 0%, #fff6fb 100%);
    transition: all .12s ease;
}
[data-testid="stFileUploader"] section:hover { border-color: #a78bfa; background: #fbf7ff; }

/* ---- Inputs / selects ---- */
[data-baseweb="select"] > div, .stNumberInput input, .stTextInput input { border-radius: 12px !important; }

/* ---- Dataframes ---- */
[data-testid="stDataFrame"] {
    border-radius: 16px; overflow: hidden; border: 1px solid #efe9ff;
    box-shadow: 0 6px 18px rgba(124,108,246,.10);
}

/* ---- Expanders ---- */
[data-testid="stExpander"] {
    border: 1px solid #efe9ff; border-radius: 16px; background: #ffffffcc;
    box-shadow: 0 4px 14px rgba(124,108,246,.08); overflow: hidden;
}

/* ---- Alerts + dividers ---- */
[data-testid="stAlert"] { border-radius: 14px; border: none; font-weight: 700; }
hr { border-color: #ece7fb; }
</style>
"""
st.markdown(friendly_theme, unsafe_allow_html=True)

st.markdown(
    '''
    <div class="app-hero">
        <div class="emoji">🚛</div>
        <div>
            <div class="title">Data Processing Toolbox</div>
            <div class="sub">Split, reconcile &amp; map your freight data — all in one friendly workspace.</div>
        </div>
    </div>
    ''',
    unsafe_allow_html=True
)

# ── Constants ─────────────────────────────────────────────────────────────────

STATES   = ['VIC', 'NSW', 'QLD', 'SA', 'WA', 'TAS', 'NT', 'ACT']

# The SECOND character of a Y-prefixed Zone code identifies the destination
# state, e.g. YNSYDMETRO -> N -> NSW, YVMELMETRO -> V -> VIC. The Zone is the
# reliable destination field; the Origin column holds warehouse / free-text
# notes and must not be used for state detection.
ZONE_LETTER_STATE = {
    'V': 'VIC', 'N': 'NSW', 'Q': 'QLD', 'S': 'SA',
    'W': 'WA',  'T': 'TAS', 'D': 'NT',  'A': 'ACT'
}

# For these invoices only NSW and VIC are valid destination states. Sydney metro
# (any *SYDMETRO zone) is folded into NSW. Anything that can't be resolved to one
# of these is routed to a "TBC" review sheet.
VALID_STATES = {'NSW', 'VIC'}

HEADER_COLOR = '2563EB'

VIP_SERVICE_KEYWORDS   = ['vip', 'elite']
VIP_INSTRUCTION_PHRASES = ['timeslot', 'time slot', 'delivery required', 'required on site']

# ── Helper functions ──────────────────────────────────────────────────────────

def _to_text_buffer(uploaded_file):
    """Decode an uploaded file to text, tolerating non-UTF-8 exports.

    Carrier statement .TXT and consignment .CSV files are frequently written as
    cp1252 ("ANSI") or UTF-16 rather than UTF-8, which makes pd.read_csv raise
    UnicodeDecodeError while reading the header row. Returns a StringIO that
    pandas can read. latin-1 is the last resort and can decode any byte, so this
    never raises.
    """
    try:
        uploaded_file.seek(0)
    except Exception:
        pass
    raw = uploaded_file.read()
    if isinstance(raw, str):
        return io.StringIO(raw)

    if raw.startswith(codecs.BOM_UTF8):
        text = raw[len(codecs.BOM_UTF8):].decode('utf-8')
    elif raw.startswith(codecs.BOM_UTF16_LE) or raw.startswith(codecs.BOM_UTF16_BE):
        text = raw.decode('utf-16')
    elif raw[:200].count(b'\x00') > 20:            # UTF-16 written without a BOM
        text = raw.decode('utf-16-le', errors='replace')
    else:
        text = None
        for enc in ('utf-8', 'cp1252', 'latin-1'):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:                           # belt and braces
            text = raw.decode('latin-1', errors='replace')
    return io.StringIO(text)


def clean_number(val):
    if pd.isna(val):
        return 0.0
    v = str(val).replace(',', '').strip()
    if v.endswith('-'):
        v = '-' + v[:-1]
    try:
        return float(v)
    except:
        return 0.0

def get_state(row):
    """Derive the delivery state from the Zone code (destination only).

    Superseded by assign_states() for the actual split — assign_states() adds an
    Origin fallback and 1800-reference matching that this row-only function can't
    do. Kept for any single-row callers.
    """
    zone = str(row.get('Zone', '')).strip().upper()
    if not zone or zone == 'NAN':
        return 'TBC'
    if 'SYDMETRO' in zone:
        return 'NSW'
    letter = zone[1] if len(zone) >= 2 and zone[0] == 'Y' else ''
    state  = ZONE_LETTER_STATE.get(letter)
    return state if state in VALID_STATES else 'TBC'
def _state_from_zone(zone):
    """Destination state from the Zone code (NSW/VIC only), else None.
    Y<letter><location>, *SYDMETRO folded into NSW."""
    z = str(zone).strip().upper()
    if not z or z == 'NAN':
        return None
    if 'SYDMETRO' in z:
        return 'NSW'
    letter = z[1] if len(z) >= 2 and z[0] == 'Y' else ''
    s = ZONE_LETTER_STATE.get(letter)
    return s if s in VALID_STATES else None


def _state_from_origin(origin):
    """State from the Origin free-text — ONLY the two valid buckets (NSW/VIC),
    matched on word boundaries so notes like 'MERGED SHIPMENT' (…meNT) or other
    states can't false-match."""
    o = str(origin).upper()
    for s in ('NSW', 'VIC'):
        if re.search(r'\b' + s + r'\b', o):
            return s
    return None


def assign_states(df):
    """Resolve the state for every row, in priority order:
      1) Zone code (destination state)
      2) Origin text (NSW / VIC warehouse) — for rows the Zone can't place
      3) another line sharing the same 1800 Delivery/Adjustment reference
         (e.g. a 'FUEL LEVY ERROR' adjustment inherits the state of the real
         shipment line with the same 1800 number)
      4) otherwise 'TBC'
    Returns a State Series aligned to df.index. Works across the whole frame,
    so ref-matching also resolves lines whose partner is in another file."""
    n = len(df)
    idx = df.index
    if n == 0:
        return pd.Series([], dtype=object)

    zone_ser   = df['Zone']                if 'Zone'                in df.columns else pd.Series([''] * n, index=idx)
    origin_ser = df['Origin']              if 'Origin'              in df.columns else pd.Series([''] * n, index=idx)
    da_ser     = df['Delivery/Adjustment'] if 'Delivery/Adjustment' in df.columns else pd.Series([''] * n, index=idx)

    # Pass 1 (zone) then Pass 2 (origin)
    resolved, refs_list = [], []
    for z, o, da in zip(zone_ser, origin_ser, da_ser):
        resolved.append(_state_from_zone(z) or _state_from_origin(o))
        refs_list.append(extract_all_refs(da))

    # Build 1800-ref -> state from rows already placed in a valid state
    ref_to_state = {}
    for s, refs in zip(resolved, refs_list):
        if s in VALID_STATES:
            for r in refs:
                ref_to_state.setdefault(r, s)

    # Pass 3: fill the rest by matching the 1800 reference; else TBC
    final = []
    for s, refs in zip(resolved, refs_list):
        if s in VALID_STATES:
            final.append(s)
        else:
            final.append(next((ref_to_state[r] for r in refs if r in ref_to_state), 'TBC'))
    return pd.Series(final, index=idx)

def split_customer_refs(ref_str):
    parts = re.split(r'[,&]+', str(ref_str))
    return [p.strip() for p in parts if p.strip()]

def extract_all_refs(ref_str):
    """Extract all 1800xxxxxx numbers from a reference string."""
    if pd.isna(ref_str):
        return []
    return re.findall(r'1800\d+', str(ref_str))

def is_vip_service(service_val):
    if pd.isna(service_val):
        return False
    s = str(service_val).lower()
    return any(kw in s for kw in VIP_SERVICE_KEYWORDS)

def is_vip_instruction(instr_val):
    if pd.isna(instr_val):
        return False
    s = str(instr_val).lower()
    return any(phrase in s for phrase in VIP_INSTRUCTION_PHRASES)

def parse_txt(uploaded_file):
    df_raw = pd.read_csv(_to_text_buffer(uploaded_file), sep='\t', dtype=str)
    summary_row_data = df_raw.iloc[-1]
    df = df_raw[:-1].copy()
    for col in ['Total (AUD)', 'GST (AUD)', 'Freight', 'LEVY', 'Load Qty', 'Paid Qty']:
        df[col] = df[col].apply(clean_number)
    df['Final Total'] = df['Total (AUD)'] + df['GST (AUD)']
    df['State']       = assign_states(df)
    client_total      = clean_number(summary_row_data['Total (AUD)'])
    return df, client_total

def parse_consignment_csvs(uploaded_files):
    """Parse one or more consignment CSV files and return a combined DataFrame."""
    frames = []
    for f in uploaded_files:
        try:
            frames.append(pd.read_csv(_to_text_buffer(f), dtype=str))
        except Exception as e:
            st.warning(f"⚠ Could not read {getattr(f, 'name', 'a .CSV')}: {e}")
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)

def enrich_txt_with_csv(df_txt, df_csv):
    """
    Join consignment CSV data onto TXT rows using Delivery/Adjustment → Reference 1.
    Adds columns: Service, Delivery Special Instructions, Carrier, Machship #
    """
    if df_csv.empty:
        df_txt['Service'] = None
        df_txt['Delivery Special Instructions'] = None
        df_txt['Carrier'] = None
        df_txt['Machship #'] = None
        df_txt['VIP'] = False
        df_txt['VIP Reason'] = ''
        return df_txt

    # Build lookup: only index CSV refs that exist in the TXT (ignore unrelated CSV rows)
    txt_refs = set(df_txt['Delivery/Adjustment'].dropna().str.strip())
    ref_lookup = {}
    for _, row in df_csv.iterrows():
        for ref in extract_all_refs(str(row.get('Reference 1', ''))):
            if ref in txt_refs and ref not in ref_lookup:
                ref_lookup[ref] = row

    enriched_rows = []
    for _, txt_row in df_txt.iterrows():
        delivery = str(txt_row.get('Delivery/Adjustment', '')).strip()
        csv_match = ref_lookup.get(delivery)
        if csv_match is not None:
            txt_row = txt_row.copy()
            txt_row['Service']                      = csv_match.get('Service', None)
            txt_row['Delivery Special Instructions'] = csv_match.get('Delivery Special Instructions', None)
            txt_row['Carrier']                      = csv_match.get('Carrier', None)
            txt_row['Machship #']                   = csv_match.get('Machship #', None)
        else:
            txt_row = txt_row.copy()
            txt_row['Service']                      = None
            txt_row['Delivery Special Instructions'] = None
            txt_row['Carrier']                      = None
            txt_row['Machship #']                   = None
        enriched_rows.append(txt_row)

    df_out = pd.DataFrame(enriched_rows)

    # Determine VIP flag and reason
    vip_flags, vip_reasons = [], []
    for _, row in df_out.iterrows():
        reasons = []
        if is_vip_service(row.get('Service')):
            reasons.append(f"Service: {row['Service']}")
        if is_vip_instruction(row.get('Delivery Special Instructions')):
            reasons.append(f"Instructions: {str(row['Delivery Special Instructions'])[:60]}")
        vip_flags.append(bool(reasons))
        vip_reasons.append(' | '.join(reasons))

    df_out['VIP'] = vip_flags
    df_out['VIP Reason'] = vip_reasons
    return df_out

def parse_csv(uploaded_file):
    df = pd.read_csv(_to_text_buffer(uploaded_file), dtype=str)
    for col in ['Quantity', 'Total Cubic', 'Rate Charge', 'Fuel Levy',
                'Rate Charge and Fuel Levy', 'Total Tax', 'Total']:
        df[col] = df[col].apply(clean_number)
    return df

def _populate_split_sheets(wb, df, client_total):
    """Populate `wb` (whose default active sheet has already been removed) with the
    standard split: per-state sheets, a Summary sheet, and a raw_data sheet (sorted
    by the 1800 reference with repeated references highlighted). Shared by the Tab 1
    processor and the Tab 4 master build so both produce an identical split."""
    # Determine which extra columns exist from CSV enrichment
    extra_cols = [c for c in ['Carrier', 'Machship #', 'Service',
                               'Delivery Special Instructions', 'VIP']
                  if c in df.columns]

    base_headers = [c for c in df.columns
                    if c not in ('State', 'Final Total', 'VIP', 'VIP Reason',
                                 'Carrier', 'Machship #', 'Service',
                                 'Delivery Special Instructions')]
    headers_with_state = base_headers + ['State'] + extra_cols

    for state in sorted(df['State'].unique()):
        state_df = df[df['State'] == state].copy()
        ws = wb.create_sheet(title=f"{state}")
        for col, header in enumerate(headers_with_state, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', start_color=HEADER_COLOR)
        for row_idx, (_, row) in enumerate(state_df.reindex(columns=headers_with_state).iterrows(), 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        total_row = len(state_df) + 2
        total_col = headers_with_state.index('Total (AUD)') + 1
        gst_col   = headers_with_state.index('GST (AUD)') + 1
        ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
        ws.cell(row=total_row, column=total_col,
                value=f'=SUM({get_column_letter(total_col)}2:{get_column_letter(total_col)}{total_row-1})'
                ).font = Font(bold=True)
        ws.cell(row=total_row, column=gst_col,
                value=f'=SUM({get_column_letter(gst_col)}2:{get_column_letter(gst_col)}{total_row-1})'
                ).font = Font(bold=True)

    # ── Summary sheet ──────────────────────────────────────────────────────
    ws_sum = wb.create_sheet(title='Summary', index=0)

    # Part A: State breakdown
    sum_headers = ['State', 'Rows', 'Total (AUD)', 'GST (AUD)', 'Total (inc GST)']
    for col, header in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color=HEADER_COLOR)
    data_row = 2
    for state in sorted(df['State'].unique()):
        state_df = df[df['State'] == state]
        total = state_df['Total (AUD)'].sum()
        gst   = state_df['GST (AUD)'].sum()
        ws_sum.cell(row=data_row, column=1, value=state)
        ws_sum.cell(row=data_row, column=2, value=len(state_df))
        ws_sum.cell(row=data_row, column=3, value=round(total, 2))
        ws_sum.cell(row=data_row, column=4, value=round(gst, 2))
        ws_sum.cell(row=data_row, column=5, value=round(total + gst, 2))
        data_row += 1

    grand_row = data_row
    ws_sum.cell(row=grand_row, column=1, value='TOTAL').font = Font(bold=True)
    for col in [2, 3, 4, 5]:
        ws_sum.cell(row=grand_row, column=col,
                    value=f'=SUM({get_column_letter(col)}2:{get_column_letter(col)}{grand_row-1})'
                    ).font = Font(bold=True)
    ws_sum.cell(row=grand_row+2, column=1, value='Statements Total (inc GST)').font = Font(bold=True)
    ws_sum.cell(row=grand_row+2, column=2, value=client_total)
    ws_sum.cell(row=grand_row+3, column=1, value='Check').font = Font(bold=True)
    ws_sum.cell(row=grand_row+3, column=2,
                value=f'=IF(E{grand_row}=B{grand_row+2},"✓ MATCH","⚠ DISCREPANCY")')

    # Part B: VIP breakdown per state (only if VIP column exists)
    if 'VIP' in df.columns:
        vip_headers = ['Delivery/Adjustment', 'Customer Name & Address',
                       'Service', 'Delivery Special Instructions',
                       'Total (AUD)', 'GST (AUD)', 'Total (inc GST)']
        vip_headers_filtered = [h for h in vip_headers if h in df.columns or h in ['Total (inc GST)']]
        num_cols = len(vip_headers_filtered)

        total_col_idx  = vip_headers_filtered.index('Total (AUD)') + 1
        gst_col_idx    = vip_headers_filtered.index('GST (AUD)') + 1
        incgst_col_idx = vip_headers_filtered.index('Total (inc GST)') + 1

        vip_df = df[df['VIP'] == True]
        vip_states = sorted(vip_df['State'].unique())

        vr = grand_row + 6  # start row for first state table

        for state in vip_states:
            state_vip = vip_df[vip_df['State'] == state]

            # Section heading
            heading_cell = ws_sum.cell(row=vr, column=1, value=f'VIP / Special Service Breakdown — {state}')
            heading_cell.font = Font(bold=True, size=12)
            vr += 1

            # Header row
            for col, header in enumerate(vip_headers_filtered, 1):
                cell = ws_sum.cell(row=vr, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', start_color='8B0000')
            vr += 1

            data_start = vr
            # Data rows
            for _, row in state_vip.iterrows():
                for col, header in enumerate(vip_headers_filtered, 1):
                    if header == 'Total (inc GST)':
                        ws_sum.cell(row=vr, column=col, value=round(row['Total (AUD)'] + row['GST (AUD)'], 2))
                    else:
                        ws_sum.cell(row=vr, column=col, value=row.get(header, ''))
                vr += 1

            # Totals row
            tot_total  = round(state_vip['Total (AUD)'].sum(), 2)
            tot_gst    = round(state_vip['GST (AUD)'].sum(), 2)
            tot_incgst = round(tot_total + tot_gst, 2)

            ws_sum.cell(row=vr, column=1, value='TOTAL').font = Font(bold=True)
            ws_sum.cell(row=vr, column=total_col_idx,  value=tot_total).font  = Font(bold=True)
            ws_sum.cell(row=vr, column=gst_col_idx,    value=tot_gst).font    = Font(bold=True)
            ws_sum.cell(row=vr, column=incgst_col_idx, value=tot_incgst).font = Font(bold=True)
            vr += 3  # gap before next state table

    # ── raw_data sheet: every invoice line, not split by state ─────────────
    # Placed at index 1 so it sits right after the Summary sheet.
    ws_raw = wb.create_sheet(title='raw_data', index=1)

    # Column G is the Delivery/Adjustment reference (the 1800xxxxx). Parse it to a
    # number, sort the sheet by it ascending, and highlight any reference that
    # appears more than once (e.g. a charge line plus its matching fuel-levy line).
    ref_col_name = 'Delivery/Adjustment'
    raw_df = df.reindex(columns=headers_with_state).copy()

    if ref_col_name in raw_df.columns:
        raw_df['_ref_num'] = pd.to_numeric(
            raw_df[ref_col_name].astype(str).str.extract(r'(\d+)')[0], errors='coerce')
        raw_df = raw_df.sort_values('_ref_num', kind='stable',
                                    na_position='last').reset_index(drop=True)
        dup_flags = (raw_df['_ref_num'].notna()
                     & raw_df['_ref_num'].duplicated(keep=False)).tolist()
        ref_col_idx = headers_with_state.index(ref_col_name) + 1
    else:
        raw_df['_ref_num'] = None
        dup_flags = [False] * len(raw_df)
        ref_col_idx = None

    dup_fill = PatternFill('solid', start_color='FFE699')  # amber = repeated 1800xxxxx

    # Header row
    for col, header in enumerate(headers_with_state, 1):
        cell = ws_raw.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color=HEADER_COLOR)

    # Data rows (sorted by reference; rows with a repeated reference are highlighted)
    for pos, (_, row) in enumerate(raw_df.iterrows()):
        row_idx  = pos + 2
        ref_num  = row['_ref_num']
        is_dup   = dup_flags[pos]
        for col_idx, header in enumerate(headers_with_state, 1):
            val = row[header]
            if col_idx == ref_col_idx and pd.notna(ref_num):
                val = int(ref_num)           # write the reference as a real number
            cell = ws_raw.cell(row=row_idx, column=col_idx, value=val)
            if is_dup:
                cell.fill = dup_fill

    # Grand total across all lines
    raw_total_row = len(raw_df) + 2
    raw_total_col = headers_with_state.index('Total (AUD)') + 1
    raw_gst_col   = headers_with_state.index('GST (AUD)') + 1
    ws_raw.cell(row=raw_total_row, column=1, value='TOTAL').font = Font(bold=True)
    ws_raw.cell(row=raw_total_row, column=raw_total_col,
                value=f'=SUM({get_column_letter(raw_total_col)}2:{get_column_letter(raw_total_col)}{raw_total_row-1})'
                ).font = Font(bold=True)
    ws_raw.cell(row=raw_total_row, column=raw_gst_col,
                value=f'=SUM({get_column_letter(raw_gst_col)}2:{get_column_letter(raw_gst_col)}{raw_total_row-1})'
                ).font = Font(bold=True)


def build_excel(df, client_total):
    """Build the split Excel (Tab 1): per-state sheets plus Summary and raw_data."""
    wb = Workbook()
    wb.remove(wb.active)
    _populate_split_sheets(wb, df, client_total)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ════════════════════════════════════════════════════════════════════════════
# MASTER CONSOLIDATION HELPERS  (Tab 4)
# ════════════════════════════════════════════════════════════════════════════

# Original .TXT column order — used when writing a re-uploadable master_file.txt.
TXT_COLUMNS = [
    'Vendor', 'Name', 'Truck', 'Origin', 'Date', 'Shipment',
    'Delivery/Adjustment', 'Customer Name & Address', 'Zone',
    'Abnormal Item Description', 'Load Qty', 'Paid Qty', 'UOM', 'Rate',
    'Freight', 'LEVY', 'Incidental (AUD)', 'Adjustment/Net (AUD)',
    'Total (AUD)', 'GST (AUD)'
]

# Numeric fields summed when consolidating per 1800 number.
CONSOL_NUM_COLS = ['Load Qty', 'Paid Qty', 'Rate', 'Freight', 'LEVY',
                   'Incidental (AUD)', 'Adjustment/Net (AUD)',
                   'Total (AUD)', 'GST (AUD)']

# ── Sell Fuel Surcharge (hardcoded from the client's fuel table) ──────────────
# Effective-date based: each rate applies from its date until the next one starts.
# The base rate applies to any shipment dated before the earliest effective date.
SELL_FSC_EFFECTIVE = [
    {'from': '22.03.2026', 'pct': 37.0},
    {'from': '04.05.2026', 'pct': 33.0},
    {'from': '11.05.2026', 'pct': 31.0},
    {'from': '18.05.2026', 'pct': 29.0},
]
SELL_FSC_BASE_PCT = 15.0   # the blank-date row — applies before 22.03.2026


def sell_fsc_for_date(date_obj):
    """Return the SELL fuel-surcharge fraction for a shipment date (effective-date
    based). Falls back to the base rate for dates before the earliest effective date."""
    pct = SELL_FSC_BASE_PCT
    if date_obj is not None:
        eff = sorted(((_parse_date(e['from']), e['pct']) for e in SELL_FSC_EFFECTIVE),
                     key=lambda x: (x[0] is None, x[0]))
        for d, p in eff:
            if d and date_obj >= d:
                pct = p
            else:
                break
    return pct / 100.0


def get_sell_fsc_df():
    """Read-only table of the hardcoded sell FSC for display on the tab."""
    rows = [{'Effective from': '(before 22.03.2026)', 'FSC %': SELL_FSC_BASE_PCT}]
    for e in SELL_FSC_EFFECTIVE:
        rows.append({'Effective from': e['from'], 'FSC %': e['pct']})
    return pd.DataFrame(rows)


def _machship_date(s):
    """Parse Machship dates like '20/May/2026' (and common fallbacks)."""
    for fmt in ('%d/%b/%Y', '%d/%B/%Y', '%d/%m/%Y', '%Y-%m-%d', '%d.%m.%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(str(s).strip()[:11], fmt).date()
        except Exception:
            pass
    return None


def compute_rate_card_expected(df_csv, rate_card, ref_col='Reference 1'):
    """Independently price each shipment from the rate card — the 'what we should
    have billed' figure. Priced off the Machship extract (clean pallet count / lane /
    despatch date), NOT the client txt (whose fuel-levy adjustment lines don't follow
    the band). For an intra-state METRO shipment: expected sell (ex GST) =
    band_rate(pallets) x pallets x (1 + sell-FSC(despatch date)). Interstate / regional
    / non-metro shipments can't be auto-priced → that reference is 'Manual'.
    Aggregated per 1800 reference."""
    empty = pd.DataFrame(columns=['Ref', 'Rate Card Expected (ex GST)', 'RC Coverage'])
    if df_csv is None or df_csv.empty or not rate_card or ref_col not in df_csv.columns:
        return empty
    band = rate_card.get('metro_band') or [60, 35, 30, 25]
    items_col = next((c for c in ['# of Items', 'Items', 'Quantity', 'Qty'] if c in df_csv.columns), None)
    fz = 'From Zone' if 'From Zone' in df_csv.columns else None
    tz = 'To Zone'   if 'To Zone'   in df_csv.columns else None
    dc = next((c for c in ['Despatch Date', 'Created Date', 'Completed/ETA Date'] if c in df_csv.columns), None)

    recs = []
    for _, r in df_csv.iterrows():
        m = re.findall(r'1800\d+', str(r.get(ref_col, '')))
        if not m:
            continue
        fzone = str(r.get(fz, '')).upper() if fz else ''
        tzone = str(r.get(tz, '')).upper() if tz else ''
        items = clean_number(r.get(items_col, 0)) if items_col else 0
        is_metro = ('METRO' in fzone and 'METRO' in tzone)      # intra-state metro lane
        if is_metro and items > 0:
            fsc = sell_fsc_for_date(_machship_date(r.get(dc)) if dc else None)
            exp = round(_band_rate(items, band) * items * (1 + fsc), 2)
            recs.append({'Ref': int(m[0]), 'exp': exp, 'ok': True})
        else:
            recs.append({'Ref': int(m[0]), 'exp': 0.0, 'ok': False})
    if not recs:
        return empty
    d = pd.DataFrame(recs)
    out = []
    for ref, grp in d.groupby('Ref'):
        if grp['ok'].all():
            out.append({'Ref': ref,
                        'Rate Card Expected (ex GST)': round(grp['exp'].sum(), 2),
                        'RC Coverage': 'Auto (metro pallet)'})
        else:
            out.append({'Ref': ref, 'Rate Card Expected (ex GST)': None,
                        'RC Coverage': 'Manual (regional/interstate/non-metro)'})
    return pd.DataFrame(out)


def parse_multiple_txt(uploaded_files):
    """Parse several client statement .TXT files (each with its own trailing
    summary row, e.g. a master_file.txt plus new invoices) and return
    (combined_df, summed_client_total, per_file_info)."""
    frames, total, info = [], 0.0, []
    for f in uploaded_files:
        try:
            df, ct = parse_txt(f)     # parse_txt seeks to 0 via _to_text_buffer
            frames.append(df.copy())
            total += ct
            info.append({'File': getattr(f, 'name', 'uploaded.txt'),
                         'Lines': len(df), 'Statement Total (AUD)': round(ct, 2)})
        except Exception as e:
            st.warning(f"⚠ Could not parse {getattr(f, 'name', 'a .TXT')}: {e}")
    if not frames:
        return pd.DataFrame(), 0.0, []
    combined = pd.concat(frames, ignore_index=True)
    # Re-resolve states across the COMBINED set so a 'FUEL LEVY ERROR' style line in
    # one file can inherit the state of its matching 1800 shipment line in another.
    combined['State'] = assign_states(combined)
    return combined, round(total, 2), info


def consolidate_by_ref(df):
    """Group every invoice line by the 1800 reference (Delivery/Adjustment) and
    sum the money / quantity columns, so charges and credits net off into one
    final figure per shipment."""
    if df.empty:
        return pd.DataFrame(columns=['Ref', 'State', 'Customer', 'Lines'] +
                            CONSOL_NUM_COLS + ['Final Total (inc GST)'])
    d = df.copy()
    d['Ref'] = pd.to_numeric(
        d['Delivery/Adjustment'].astype(str).str.extract(r'(\d+)')[0], errors='coerce')
    for c in CONSOL_NUM_COLS:
        d[c] = d[c].apply(clean_number) if c in d.columns else 0.0

    g       = d.groupby('Ref', dropna=False)
    grouped = g[CONSOL_NUM_COLS].sum()
    lines   = g.size().rename('Lines')
    state   = g['State'].agg(
        lambda s: s.dropna().iloc[0] if s.notna().any() else 'TBC').rename('State')
    cust    = g['Customer Name & Address'].agg(
        lambda s: s.dropna().astype(str).iloc[0] if s.notna().any() else '').rename('Customer')

    out = pd.concat([lines, state, cust, grouped], axis=1).reset_index()
    out['Final Total (inc GST)'] = out['Total (AUD)'] + out['GST (AUD)']
    for c in CONSOL_NUM_COLS + ['Final Total (inc GST)']:
        out[c] = out[c].round(2)
    out = out.sort_values('Ref', na_position='last').reset_index(drop=True)
    cols = ['Ref', 'State', 'Customer', 'Lines'] + CONSOL_NUM_COLS + ['Final Total (inc GST)']
    return out[cols]


def detect_ref_column(df_csv):
    """Best-guess which Machship CSV column holds the 1800 reference."""
    for p in ['Reference 1', 'Reference', 'Ref 1', 'Ref',
              'Customer Reference', 'Con Note', 'Connote']:
        if p in df_csv.columns:
            return p
    best, best_hits = None, 0
    for c in df_csv.columns:
        hits = df_csv[c].astype(str).str.contains(r'1800\d{4,}', regex=True, na=False).sum()
        if hits > best_hits:
            best, best_hits = c, hits
    return best if best_hits else (df_csv.columns[0] if len(df_csv.columns) else None)


def detect_sell_column(df_csv):
    """Best-guess the Machship 'sell' (ex GST) amount column. 'Sell Ex Tax' is the
    real ex-GST sell in the Machship export (Base Sell + Fuel Sell); 'Total Sell'
    is inc GST, so it's only a last-resort fallback."""
    for p in ['Sell Ex Tax', 'Total Sell Ex GST', 'Sell Ex GST', 'Sell (Ex GST)',
              'Rate Charge and Fuel Levy', 'Total Sell', 'Sell', 'Total']:
        if p in df_csv.columns:
            return p
    return None


def machship_by_ref(df_csv, ref_col, sell_col):
    """Sum the Machship sell amount (ex GST) per 1800 reference.
    Returns columns: Ref, Machship Sell (ex GST), Machship Rows."""
    empty = pd.DataFrame(columns=['Ref', 'Machship Sell (ex GST)', 'Machship Rows'])
    if df_csv is None or df_csv.empty or not ref_col or not sell_col:
        return empty
    rows = []
    for _, r in df_csv.iterrows():
        refs = re.findall(r'1800\d+', str(r.get(ref_col, '')))
        if not refs:
            continue
        rows.append({'Ref': int(refs[0]), 'Sell': clean_number(r.get(sell_col, 0))})
    if not rows:
        return empty
    md = pd.DataFrame(rows)
    out = md.groupby('Ref').agg(
        **{'Machship Sell (ex GST)': ('Sell', 'sum'),
           'Machship Rows': ('Sell', 'size')}).reset_index()
    out['Machship Sell (ex GST)'] = out['Machship Sell (ex GST)'].round(2)
    return out


def build_discrepancy(consol_df, mship_df, rc_df=None):
    """Per 1800 reference: client paid (Total AUD ex GST) vs Machship sell (ex GST),
    and — when a rate card is supplied — the rate-card 'should have billed' (ex GST)
    as the source of truth. Status reflects client vs Machship."""
    client_refs = set(consol_df['Ref']) if not consol_df.empty else set()
    mship_refs  = set(mship_df['Ref'])  if not mship_df.empty  else set()
    if not client_refs and not mship_refs:
        return pd.DataFrame()

    left = (consol_df[['Ref', 'State', 'Customer', 'Total (AUD)']]
            .rename(columns={'Total (AUD)': 'Client Paid (ex GST)'})
            if not consol_df.empty else
            pd.DataFrame(columns=['Ref', 'State', 'Customer', 'Client Paid (ex GST)']))
    right = mship_df if not mship_df.empty else \
            pd.DataFrame(columns=['Ref', 'Machship Sell (ex GST)', 'Machship Rows'])

    m = left.merge(right, on='Ref', how='outer')
    m['Client Paid (ex GST)']   = m['Client Paid (ex GST)'].fillna(0).round(2)
    m['Machship Sell (ex GST)'] = m['Machship Sell (ex GST)'].fillna(0).round(2)
    m['Machship Rows']          = m['Machship Rows'].fillna(0).astype(int)
    m['State']    = m['State'].fillna('TBC')
    m['Customer'] = m['Customer'].fillna('')
    m['Difference (Client - Machship)'] = (m['Client Paid (ex GST)']
                                           - m['Machship Sell (ex GST)']).round(2)

    def status(ref, diff):
        inc, inm = ref in client_refs, ref in mship_refs
        if inc and not inm: return '⚠ No Machship data'
        if inm and not inc: return '⚠ No client line'
        if abs(diff) <= 0.01: return '✓ Match'
        return '❌ Discrepancy'
    m['Status'] = [status(ref, diff) for ref, diff
                   in zip(m['Ref'], m['Difference (Client - Machship)'])]

    cols = ['Ref', 'State', 'Customer', 'Client Paid (ex GST)', 'Machship Sell (ex GST)',
            'Difference (Client - Machship)', 'Machship Rows', 'Status']

    # Rate card = independent source of truth (when supplied)
    if rc_df is not None and not rc_df.empty:
        m = m.merge(rc_df, on='Ref', how='left')
        m['RC Coverage'] = m['RC Coverage'].fillna('Manual (regional/interstate/tonne)')
        m['Machship vs Rate Card'] = (m['Machship Sell (ex GST)']
                                      - m['Rate Card Expected (ex GST)']).round(2)
        m['Client vs Rate Card'] = (m['Client Paid (ex GST)']
                                    - m['Rate Card Expected (ex GST)']).round(2)
        cols = ['Ref', 'State', 'Customer', 'Client Paid (ex GST)', 'Machship Sell (ex GST)',
                'Rate Card Expected (ex GST)', 'Difference (Client - Machship)',
                'Machship vs Rate Card', 'Client vs Rate Card', 'RC Coverage',
                'Machship Rows', 'Status']

    m = m.sort_values('Ref', na_position='last').reset_index(drop=True)
    return m[[c for c in cols if c in m.columns]]


def build_state_comparison(consol_df, mship_df, rc_df=None):
    """Per state: 'We Thought' = Machship sell (ex GST) vs 'Their Figure' = client
    Total (AUD) ex GST, plus (when supplied) 'Rate Card Should Be' = rate-card
    expected (ex GST). Machship / rate-card refs are attributed to the state of the
    matching client line; Machship refs with no client line fall under 'Unmatched'."""
    if not consol_df.empty:
        client    = consol_df.groupby('State')['Total (AUD)'].sum().rename('Their Figure (ex GST)')
        ref_state = dict(zip(consol_df['Ref'], consol_df['State']))
    else:
        client, ref_state = pd.Series(dtype=float, name='Their Figure (ex GST)'), {}

    if not mship_df.empty:
        tmp = mship_df.copy()
        tmp['State'] = tmp['Ref'].map(ref_state).fillna('Unmatched')
        mship = tmp.groupby('State')['Machship Sell (ex GST)'].sum().rename('We Thought (ex GST)')
    else:
        mship = pd.Series(dtype=float, name='We Thought (ex GST)')

    if client.empty and mship.empty:
        return pd.DataFrame()

    parts = [mship, client]
    have_rc = rc_df is not None and not rc_df.empty
    if have_rc:
        tmp = rc_df.dropna(subset=['Rate Card Expected (ex GST)']).copy()
        tmp['State'] = tmp['Ref'].map(ref_state).fillna('Unmatched')
        rc = tmp.groupby('State')['Rate Card Expected (ex GST)'].sum().rename('Rate Card Should Be (ex GST)')
        parts.append(rc)

    out = pd.concat(parts, axis=1).fillna(0.0)
    out['Difference (Client - Machship)'] = (out['Their Figure (ex GST)']
                                             - out['We Thought (ex GST)']).round(2)
    for c in ['We Thought (ex GST)', 'Their Figure (ex GST)', 'Rate Card Should Be (ex GST)']:
        if c in out.columns:
            out[c] = out[c].round(2)
    out = out.reset_index().rename(columns={'index': 'State'})
    out['_o'] = out['State'].apply(lambda s: 1 if s == 'Unmatched' else 0)
    out = out.sort_values(['_o', 'State']).drop(columns='_o').reset_index(drop=True)

    order = ['State', 'We Thought (ex GST)', 'Their Figure (ex GST)']
    if have_rc:
        order.append('Rate Card Should Be (ex GST)')
    order.append('Difference (Client - Machship)')
    return out[[c for c in order if c in out.columns]]


def build_master_txt(df):
    """Serialise the combined client data back to a tab-separated master_file.txt
    (original .TXT columns + a trailing summary row) so it re-uploads cleanly and
    keeps growing next time. The summary row's Total (AUD) holds the INC-GST grand
    total, matching the source statements (that's the figure parse_txt reads).
    Written as UTF-8; _to_text_buffer reads it back regardless of encoding."""
    out = df.reindex(columns=TXT_COLUMNS).copy()
    ex_gst  = pd.to_numeric(out['Total (AUD)'], errors='coerce').fillna(0).sum()
    gst     = pd.to_numeric(out['GST (AUD)'],   errors='coerce').fillna(0).sum()
    inc_gst = round(ex_gst + gst, 2)
    summary = {c: '' for c in TXT_COLUMNS}
    summary['Vendor'] = 'TOTAL'
    summary['Total (AUD)'] = inc_gst
    out = pd.concat([out, pd.DataFrame([summary])], ignore_index=True)
    return out.to_csv(sep='\t', index=False).encode('utf-8')


def _py(v):
    """Convert numpy scalars to plain Python types for openpyxl."""
    if hasattr(v, 'item'):
        try: return v.item()
        except Exception: return v
    return v


def _write_df_sheet(ws, df, title, row_fill_fn=None, total_cols=None):
    """Write a DataFrame to a sheet: title (row 1), header (row 2), data (row 3+),
    optional per-row fill and an optional TOTAL row summing `total_cols`."""
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    cols = list(df.columns)
    for c, col in enumerate(cols, 1):
        cell = ws.cell(row=2, column=c, value=col)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color=HEADER_COLOR)
    r = 3
    for _, row in df.iterrows():
        fill = row_fill_fn(row) if row_fill_fn else None
        for c, col in enumerate(cols, 1):
            val = _py(row[col])
            if isinstance(val, float):
                val = round(val, 2)
            cell = ws.cell(row=r, column=c, value=val)
            if fill is not None:
                cell.fill = fill
        r += 1
    if total_cols:
        ws.cell(row=r, column=1, value='TOTAL').font = Font(bold=True)
        for c, col in enumerate(cols, 1):
            if col in total_cols and col in df.columns:
                tot = round(pd.to_numeric(df[col], errors='coerce').fillna(0).sum(), 2)
                ws.cell(row=r, column=c, value=tot).font = Font(bold=True)
    return ws


def build_master_excel(df_all, client_total, consol_df, mship_df, disc_df, state_cmp_df):
    """Build the master workbook: the standard split (Summary, raw_data, states)
    plus consolidated_by_1800, discrepancies, and state_vs_machship sheets."""
    wb = Workbook()
    wb.remove(wb.active)
    _populate_split_sheets(wb, df_all, client_total)

    red   = PatternFill('solid', start_color='C00000')
    amber = PatternFill('solid', start_color='B8860B')
    green = PatternFill('solid', start_color='1F7A1F')
    blue  = PatternFill('solid', start_color='2F5597')

    # consolidated_by_1800 — flag net-credit shipments (negative Total) in blue
    def consol_fill(row):
        try:
            return blue if float(row['Total (AUD)']) < 0 else None
        except Exception:
            return None
    _write_df_sheet(
        wb.create_sheet('consolidated_by_1800'), consol_df,
        'Consolidated cost per 1800 number (charges + credits netted)',
        row_fill_fn=consol_fill,
        total_cols=CONSOL_NUM_COLS + ['Final Total (inc GST)'])

    # discrepancies — colour by status
    def disc_fill(row):
        s = row.get('Status', '')
        if s == '❌ Discrepancy': return red
        if s == '✓ Match':        return green
        if isinstance(s, str) and s.startswith('⚠'): return amber
        return None
    if not disc_df.empty:
        _write_df_sheet(
            wb.create_sheet('discrepancies'), disc_df,
            'Client paid vs Machship sell vs Rate card, per 1800 number (ex GST)',
            row_fill_fn=disc_fill,
            total_cols=['Client Paid (ex GST)', 'Machship Sell (ex GST)',
                        'Rate Card Expected (ex GST)', 'Difference (Client - Machship)',
                        'Machship vs Rate Card', 'Client vs Rate Card'])

    # state_vs_machship
    if not state_cmp_df.empty:
        _write_df_sheet(
            wb.create_sheet('state_vs_machship'), state_cmp_df,
            'Per state: what we thought we make vs their figure (ex GST)',
            total_cols=['We Thought (ex GST)', 'Their Figure (ex GST)',
                        'Rate Card Should Be (ex GST)', 'Difference (Client - Machship)'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ════════════════════════════════════════════════════════════════════════════
# RATE-CARD RECONCILIATION HELPERS  (Tab 2)
# ════════════════════════════════════════════════════════════════════════════
#
# ONLINE / STREAMLIT-CLOUD BUILD
# ------------------------------
# This version stores NOTHING. There is no database. The rate card and all
# invoice/CSV files must be uploaded every session and live only in memory for
# the duration of the run. Nothing is written to disk.
#
# The only hardcoded values are the Fuel Surcharge (FSC) rates below — these
# are not confidential, so they are kept in code for convenience. Edit them
# here whenever the FSC changes.

ZONE_PREFIX_STATE = {
    'V': 'VIC', 'N': 'NSW', 'S': 'SA', 'Q': 'QLD',
    'W': 'WA',  'T': 'TAS', 'D': 'NT', 'A': 'ACT'
}

# ── Fuel Surcharge (FSC) config — the ONLY hardcoded metric ───────────────────
# CURRENT_FSC_PCT is the default applied to every invoice line whose date is not
# covered by a date-range override below. Update it when the current week's FSC
# changes. Date ranges use DD.MM.YYYY.

CURRENT_FSC_PCT = 29.0

FSC_OVERRIDES = [
    {"from": "01.05.2026", "to": "01.05.2026", "pct": 42.0, "note": "1 May"},
    {"from": "04.05.2026", "to": "08.05.2026", "pct": 33.0, "note": "4-8 May"},
    {"from": "11.05.2026", "to": "19.05.2026", "pct": 31.0, "note": "11-19 May"},
    {"from": "20.05.2026", "to": "31.05.2026", "pct": 29.0, "note": "20-31 May"},
]

def get_fsc_overrides_df():
    """Build the FSC override table (in memory) from the hardcoded config."""
    if not FSC_OVERRIDES:
        return pd.DataFrame(columns=['date_from', 'date_to', 'fsc_pct', 'note'])
    return pd.DataFrame([
        {'date_from': o['from'], 'date_to': o['to'], 'fsc_pct': o['pct'], 'note': o.get('note', '')}
        for o in FSC_OVERRIDES
    ])


# ── Rate-card parsing ─────────────────────────────────────────────────────────

def parse_rate_card(file_like):
    """Parse the Rohlig/Sugar transportation rate card → structured dict."""
    wb = load_workbook(file_like, data_only=True)
    parsed = {'effective': None, 'zone_bands': {}, 'metro_band': None,
              'cubic_factor': None, 'basic_charge_kg': None}

    if 'Transportation - Metro,Regional' not in wb.sheetnames:
        return parsed

    ws = wb['Transportation - Metro,Regional']
    eff = ws.cell(row=1, column=2).value or ''
    m = re.search(r'(\d{2}\.\d{2}\.\d{4})', str(eff))
    parsed['effective'] = m.group(1) if m else None

    cf = ws.cell(row=6, column=4).value
    if isinstance(cf, (int, float)):
        parsed['cubic_factor'] = cf

    note = ws.cell(row=3, column=32).value or ''
    bc = re.search(r'\$?\s*([\d.]+)', str(note))
    if bc:
        try: parsed['basic_charge_kg'] = float(bc.group(1))
        except: pass

    # Regional pallet block: cols J-R (10-18); zone=col 13, bands=cols 15-18
    zone_bands = {}
    for r in range(14, ws.max_row + 1):
        zone = ws.cell(row=r, column=13).value
        p1, p25, p611, p12 = (ws.cell(row=r, column=col).value for col in (15, 16, 17, 18))
        if zone and isinstance(p1, (int, float)):
            zone_bands.setdefault(str(zone).strip(), [p1, p25, p611, p12])
    parsed['zone_bands'] = zone_bands

    if zone_bands:
        band_counts = Counter(tuple(v) for v in zone_bands.values())
        parsed['metro_band'] = list(band_counts.most_common(1)[0][0])

    return parsed

# ── Reconciliation logic ────────────────────────────────────────────────────

def _origin_state(origin_text, zone):
    for s in STATES:
        if s in str(origin_text):
            return s
    z = str(zone)
    if len(z) >= 2:
        return ZONE_PREFIX_STATE.get(z[1].upper(), 'UNKNOWN')
    return 'UNKNOWN'

def _dest_state(zone):
    z = str(zone)
    if len(z) >= 2:
        return ZONE_PREFIX_STATE.get(z[1].upper(), 'UNKNOWN')
    return 'UNKNOWN'

def _band_rate(qty, band):
    q = round(qty)
    if q <= 1:  return band[0]
    if q <= 5:  return band[1]
    if q <= 11: return band[2]
    return band[3]

def _parse_date(d):
    """TXT dates are DD.MM.YYYY."""
    for fmt in ('%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(str(d).strip(), fmt).date()
        except: pass
    return None

def fsc_for_date(date_obj, fsc_df, default_fsc=None):
    """
    Return (fsc_fraction, source) for a given date.
    Date-range entries take priority; otherwise fall back to the current default FSC.
    default_fsc is a percentage (e.g. 31.0) or None.
    """
    if date_obj is not None and not fsc_df.empty:
        for _, r in fsc_df.iterrows():
            df_ = _parse_date(r['date_from']); dt_ = _parse_date(r['date_to'])
            if df_ and dt_ and df_ <= date_obj <= dt_:
                return float(r['fsc_pct']) / 100.0, 'range'
    if default_fsc is not None:
        return float(default_fsc) / 100.0, 'default'
    return None, None

def reconcile_txt(df_txt, rate_card, fsc_df, default_fsc=None):
    """
    Compare each TXT line against the rate card.
    Auto-verify same-state metro PAL lines; flag interstate/regional/tonne for manual review.
    default_fsc (percentage) is applied to any line whose date isn't covered by a date-range entry.
    """
    band = rate_card.get('metro_band') or [60, 35, 30, 25]
    records = []
    for _, row in df_txt.iterrows():
        zone        = row.get('Zone', '')
        uom         = str(row.get('UOM', '')).strip()
        qty         = clean_number(row.get('Paid Qty', 0))
        chg_rate    = clean_number(row.get('Rate', 0))
        chg_freight = clean_number(row.get('Freight', 0))
        chg_levy    = clean_number(row.get('LEVY', 0))
        date_obj    = _parse_date(row.get('Date'))
        o_state     = _origin_state(row.get('Origin', ''), zone)
        d_state     = _dest_state(zone)
        fsc, fsc_src = fsc_for_date(date_obj, fsc_df, default_fsc)

        is_metro_pal = ('METRO' in str(zone).upper()) and uom == 'PAL' and (o_state == d_state)

        rec = {
            'Date': row.get('Date'),
            'Delivery/Adjustment': row.get('Delivery/Adjustment'),
            'Customer Name & Address': row.get('Customer Name & Address'),
            'Zone': zone, 'UOM': uom, 'Qty': qty,
            'Charged Rate': chg_rate, 'Charged Freight': round(chg_freight, 2),
            'Charged LEVY': round(chg_levy, 2),
            'Charged Total': round(chg_freight + chg_levy, 2),
        }

        if is_metro_pal:
            exp_rate    = _band_rate(qty, band)
            exp_freight = round(exp_rate * qty, 2)
            if fsc is not None:
                exp_levy  = round(exp_freight * fsc, 2)
                exp_total = round(exp_freight + exp_levy, 2)
                levy_diff = round(chg_levy - exp_levy, 2)
            else:
                exp_levy = exp_total = levy_diff = None
            rec.update({
                'Check': 'Auto (metro pallet)',
                'Expected Rate': exp_rate,
                'Expected Freight': exp_freight,
                'FSC %': round(fsc * 100, 2) if fsc is not None else None,
                'FSC Source': fsc_src,
                'Expected LEVY': exp_levy,
                'Expected Total': exp_total,
                'Rate Diff': round(chg_rate - exp_rate, 2),
                'Freight Diff': round(chg_freight - exp_freight, 2),
                'LEVY Diff': levy_diff,
                'Total Diff': round(chg_freight + chg_levy - exp_total, 2) if exp_total is not None else None,
            })
            ft_diff = abs(rec['Freight Diff']) > 0.01
            lv_diff = (levy_diff is not None and abs(levy_diff) > 0.01)
            if rec['FSC %'] is None:
                rec['Status'] = '⚠ No FSC for date'
            elif ft_diff or lv_diff:
                rec['Status'] = '❌ Mismatch'
            else:
                rec['Status'] = '✓ OK'
        else:
            reason = []
            if o_state != d_state: reason.append(f'interstate {o_state}→{d_state}')
            if uom != 'PAL':       reason.append(f'UOM={uom}')
            if 'METRO' not in str(zone).upper(): reason.append('non-metro zone')
            rec.update({
                'Check': 'Manual review (' + ', '.join(reason) + ')',
                'Expected Rate': None, 'Expected Freight': None,
                'FSC %': round(fsc * 100, 2) if fsc is not None else None,
                'FSC Source': fsc_src,
                'Expected LEVY': None, 'Expected Total': None,
                'Rate Diff': None, 'Freight Diff': None, 'LEVY Diff': None, 'Total Diff': None,
                'Status': '🔍 Manual review',
            })
        records.append(rec)

    return pd.DataFrame(records)

def reconcile_csv(df_csv, recon_txt_df):
    """
    Compare Machship CSV 'Total Sell' against the rate-card expected total — but ONLY
    for shipments that are on the invoice (i.e. whose delivery ref appears on a TXT line).
    CSV rows not on the invoice are ignored entirely.
    """
    if df_csv.empty:
        return pd.DataFrame()

    # 1) Invoice delivery refs — every ref that appears on a TXT line (the invoice).
    invoice_refs = set()
    for _, r in recon_txt_df.iterrows():
        ref = str(r.get('Delivery/Adjustment', '')).strip()
        if ref:
            invoice_refs.add(ref)

    # 2) Expected total per ref (only auto-checked metro lines have one).
    exp_by_ref = {}
    for _, r in recon_txt_df.iterrows():
        ref = str(r.get('Delivery/Adjustment', '')).strip()
        if ref and pd.notna(r.get('Expected Total')):
            exp_by_ref[ref] = r['Expected Total']

    rows = []
    for _, row in df_csv.iterrows():
        refs = re.findall(r'1800\d+', str(row.get('Reference 1', '')))
        # Only keep CSV rows that belong to the invoice.
        invoice_match = next((ref for ref in refs if ref in invoice_refs), None)
        if invoice_match is None:
            continue

        total_sell = clean_number(row.get('Total Sell', 0))
        matched_exp = next((exp_by_ref[ref] for ref in refs if ref in exp_by_ref), None)
        has_exp = matched_exp is not None and pd.notna(matched_exp)
        rows.append({
            'Machship #': row.get('Machship #'),
            'Reference 1': row.get('Reference 1'),
            'To Name': row.get('To Name'),
            'CSV Total Sell': round(total_sell, 2),
            'Rate Card Expected': round(matched_exp, 2) if has_exp else None,
            'Diff': round(total_sell - matched_exp, 2) if has_exp else None,
            'Status': ('✓ OK' if (has_exp and abs(total_sell - matched_exp) <= 0.01)
                       else ('❌ Mismatch' if has_exp else '🔍 Manual review')),
        })
    return pd.DataFrame(rows)

# ════════════════════════════════════════════════════════════════════════════
# TGE MAPPER HELPERS  (Tab 5)
# ════════════════════════════════════════════════════════════════════════════
#
# Reads the line items out of TGE invoice PDFs and maps each invoice shipment to
# the Machship consignment export, pulling on the Machship # and Reference 2 (the
# W-number). Matching key: invoice "Customer Reference 1" → consignment
# "Reference 1", with a Machship-# fallback (the customer often enters the
# Machship number itself as their reference). Consignments that aren't on the
# invoice are ignored. Invoice lines that can't be matched are left blank and
# highlighted orange for manual review.

# Invoice line-item columns, in the order they appear on the TGE PDF.
INVOICE_COLUMNS = [
    'Trading Account Number', 'Site Number', 'Date', 'Team GE Reference',
    'Customer Reference 1', 'Product', 'Service', 'Origin', 'Destination',
    'Qty', 'Cubic', 'Declared kg', 'Charged kg', 'Freight Charge', 'Fuel Charge',
    'Other Charges', 'Total Charge Excl. GST', 'GST', 'Total Charge Incl. GST'
]

TGE_UNMATCHED_FILL = 'FFC000'   # orange for Excel unmatched rows


def _norm_ref(s):
    """Normalise a reference for matching: collapse whitespace, upper-case,
    treat NaN/None as empty."""
    if s is None:
        return ''
    txt = re.sub(r'\s+', ' ', str(s)).strip()
    return '' if txt.lower() == 'nan' else txt.upper()


def _ref_tokens(s):
    """Split a (possibly multi-value) reference into individual tokens.
    Handles separators like ',', '&', '/', and ' - '."""
    raw = re.split(r'[,&/]|\s-\s|\s+', str(s if s is not None else ''))
    return [_norm_ref(p) for p in raw if _norm_ref(p)]


def _clean_cell(v):
    """Clean a raw PDF/CSV cell → single-line trimmed string ('' for NaN/None)."""
    if v is None:
        return ''
    s = re.sub(r'\s+', ' ', str(v)).strip()
    return '' if s.lower() == 'nan' else s


def parse_invoice_pdf(file_like, source_name=''):
    """Extract line items from a TGE invoice PDF.

    Returns (rows, invoice_number). Each row is a dict keyed by the invoice
    column names plus 'Invoice #'. Uses pdfplumber's table extraction, locating
    the 'Team GE Reference / Customer Reference 1' header on each page."""
    import pdfplumber   # lazy import so the rest of the app runs without it
    rows, invoice_no = [], None
    try:
        file_like.seek(0)
    except Exception:
        pass
    with pdfplumber.open(file_like) as pdf:
        for page in pdf.pages:
            if invoice_no is None:
                txt = page.extract_text() or ''
                m = re.search(r'Invoice Number:?\s*([0-9]+)', txt)
                if m:
                    invoice_no = m.group(1)
            for t in page.extract_tables():
                if not t:
                    continue
                hdr_idx = None
                for i, r in enumerate(t):
                    joined = ' '.join(str(c) for c in r if c)
                    if 'Team GE' in joined and 'Customer' in joined:
                        hdr_idx = i
                        break
                if hdr_idx is None:
                    continue
                header = [_clean_cell(c) for c in t[hdr_idx]]
                for r in t[hdr_idx + 1:]:
                    d = {}
                    for k in range(len(header)):
                        key = header[k]
                        if not key:
                            continue
                        d[key] = _clean_cell(r[k]) if k < len(r) else ''
                    teamge  = d.get('Team GE Reference', '')
                    tot     = d.get('Total Charge Incl. GST', '')
                    custref = d.get('Customer Reference 1', '')
                    joined  = ' '.join(v for v in d.values() if v)
                    if not teamge and not tot and not custref:
                        continue
                    if joined.startswith('Sub Total') or 'Total Payable' in joined:
                        continue
                    d['Invoice #'] = invoice_no or source_name
                    rows.append(d)
    return rows, (invoice_no or source_name)


def build_consignment_lookup(df_csv):
    """Build lookups from the consignment export for matching invoice refs:
    Reference 1 (exact) → rec, Reference 1 tokens → rec, Machship # (exact) → rec.
    Each rec holds the Machship # and Reference 2 to copy onto the invoice line."""
    ref_full, ref_tok, ms_full = {}, {}, {}
    if df_csv is None or len(df_csv) == 0:
        return ref_full, ref_tok, ms_full
    r1col = 'Reference 1' if 'Reference 1' in df_csv.columns else None
    r2col = 'Reference 2' if 'Reference 2' in df_csv.columns else None
    mscol = 'Machship #'  if 'Machship #'  in df_csv.columns else None
    for _, row in df_csv.iterrows():
        rec = {'Machship #':  _clean_cell(row.get(mscol, '')) if mscol else '',
               'Reference 2': _clean_cell(row.get(r2col, '')) if r2col else ''}
        if r1col:
            n = _norm_ref(row.get(r1col, ''))
            if n and n not in ref_full:
                ref_full[n] = rec
            for tk in _ref_tokens(row.get(r1col, '')):
                if len(tk) >= 4 and re.search(r'\d', tk) and tk not in ref_tok:
                    ref_tok[tk] = rec
        if mscol:
            mn = _norm_ref(row.get(mscol, ''))
            if mn and mn not in ms_full:
                ms_full[mn] = rec
    return ref_full, ref_tok, ms_full


def map_invoice_row(custref, lookups):
    """Resolve a single invoice 'Customer Reference 1' to a consignment record.
    Priority: exact Reference 1 → partial Reference 1 token → Machship #.
    Returns (rec_or_None, how)."""
    ref_full, ref_tok, ms_full = lookups
    n = _norm_ref(custref)
    if not n:
        return None, 'Unmatched'
    if n in ref_full:
        return ref_full[n], 'Reference 1'
    for tk in _ref_tokens(custref):
        if len(tk) >= 4 and re.search(r'\d', tk) and tk in ref_tok:
            return ref_tok[tk], 'Reference 1 (partial)'
    if n in ms_full:
        return ms_full[n], 'Machship #'
    return None, 'Unmatched'


def build_tge_mapping(invoice_rows, df_csv):
    """Map every invoice line to the consignment export, adding Machship #,
    Reference 2, Matched On and a boolean _matched flag."""
    lookups = build_consignment_lookup(df_csv)
    out = []
    for d in invoice_rows:
        rec = dict(d)
        m, how = map_invoice_row(d.get('Customer Reference 1', ''), lookups)
        rec['Machship #']  = m['Machship #']  if m else ''
        rec['Reference 2'] = m['Reference 2'] if m else ''
        rec['Matched On']  = how
        rec['_matched']    = m is not None
        out.append(rec)
    return pd.DataFrame(out)


def tge_display_order(df):
    """Column order for the mapped table: Customer Reference 1, then the two added
    consignment columns (Machship #, Reference 2), then the remaining invoice
    columns. Only keeps columns present; the 'Matched On' column is dropped."""
    lead  = ['Customer Reference 1', 'Machship #', 'Reference 2']
    rest  = ['Invoice #'] + [c for c in INVOICE_COLUMNS if c != 'Customer Reference 1']
    order = lead + [c for c in rest if c not in lead]
    return [c for c in order if c in df.columns]


# Columns for the second, cut-down summary sheet.
TGE_SUMMARY_COLUMNS = ['Customer Reference 1', 'Machship #', 'Reference 2', 'Team GE Reference']


def build_tge_mapper_excel(df, order):
    """Write the mapped invoice→consignment data to Excel with two sheets:
      1) 'Invoice to Consignment' — every invoice column + Machship #/Reference 2.
      2) 'Reference Summary' — only Customer Reference 1, Machship #, Reference 2
         and Team GE Reference.
    On both sheets, unmatched invoice rows (blank Machship #/Reference 2) are
    highlighted orange."""
    wb = Workbook()
    wb.remove(wb.active)
    orange = PatternFill('solid', start_color=TGE_UNMATCHED_FILL)
    hdr    = PatternFill('solid', start_color=HEADER_COLOR)

    def _write_sheet(ws, cols, title):
        cols = [c for c in cols if c in df.columns]
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
        for c, col in enumerate(cols, 1):
            cell = ws.cell(row=2, column=c, value=col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = hdr
        for r, (_, row) in enumerate(df.iterrows(), 3):
            unmatched = not bool(row.get('_matched', True))
            for c, col in enumerate(cols, 1):
                cell = ws.cell(row=r, column=c, value=_clean_cell(row.get(col, '')))
                if unmatched:
                    cell.fill = orange
        ws.freeze_panes = 'A3'
        for c, col in enumerate(cols, 1):
            try:
                sample = [len(str(col))] + [len(str(v)) for v in df[col].astype(str).head(200)]
                ws.column_dimensions[get_column_letter(c)].width = max(10, min(42, max(sample) + 2))
            except Exception:
                ws.column_dimensions[get_column_letter(c)].width = 16

    _write_sheet(wb.create_sheet('Invoice to Consignment'),
                 [c for c in order if c in df.columns],
                 'Invoice lines mapped to Machship consignment export')
    _write_sheet(wb.create_sheet('Reference Summary'),
                 TGE_SUMMARY_COLUMNS,
                 'Key references (Customer Ref 1 · Machship # · Reference 2 · Team GE Ref)')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Tabs ──────────────────────────────────────────────────────────────────────

tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📊 RCTI Processor",
    "🔍 Invoice Reconciliation",
    "📸 Screenshot to Excel (Beta)",
    "🗂 Master Consolidation",
    "🗺️ TGE Mapper",
])

# ════════════════════════════════════════════════════════════════════════════
# TAB 1
# ════════════════════════════════════════════════════════════════════════════

with tab1:
    st.write("Upload your `.TXT` statement file and optionally one or more consignment `.CSV` files to enrich with shipment data.")

    col_up1, col_up2 = st.columns(2)
    with col_up1:
        uploaded_txt = st.file_uploader("Statement (.TXT)", type=['txt', 'TXT'], key="tab1_txt")
    with col_up2:
        uploaded_csvs = st.file_uploader(
            "Consignment report(s) (.CSV) — optional, select multiple if needed",
            type=['csv', 'CSV'], key="tab1_csvs", accept_multiple_files=True
        )

    if uploaded_txt:
        if uploaded_txt.size > 20 * 1024 * 1024:
            st.error("⚠ File exceeds the 20MB limit.")
            st.stop()
        with st.spinner("Processing..."):
            df, client_total = parse_txt(uploaded_txt)

            # Enrich with CSV if provided
            if uploaded_csvs:
                df_csv_combined = parse_consignment_csvs(uploaded_csvs)
                df = enrich_txt_with_csv(df, df_csv_combined)
                csv_matched = int(df['Machship #'].notna().sum())
                total = len(df)
                if csv_matched == total:
                    st.success(f"✓ {csv_matched}/{total} shipments on TXT matched with consignment data.")
                else:
                    st.warning(f"⚠ {csv_matched}/{total} shipments on TXT matched with consignment data.")
                    unmatched_df = df[df['Machship #'].isna()]
                    with st.expander(f"Show {len(unmatched_df)} TXT shipment(s) still needing consignment data"):
                        st.dataframe(
                            unmatched_df[['Date', 'Truck', 'State', 'Delivery/Adjustment', 'Customer Name & Address']],
                            use_container_width=True, hide_index=True
                        )
            else:
                df['VIP'] = False
                df['VIP Reason'] = ''

            our_total   = df['Total (AUD)'].sum()
            our_gst     = df['GST (AUD)'].sum()
            our_inc_gst = our_total + our_gst
            match       = round(our_inc_gst, 2) == round(client_total, 2)
            vendor      = df['Vendor'].iloc[0]

        st.divider()
        col1, col2, col3 = st.columns(3)
        col1.metric("Invoice Lines", len(df))
        col2.metric("Trucks", df['Truck'].nunique())
        col3.metric("States", df['State'].nunique())

        st.divider()
        st.subheader("Totals")
        col1, col2, col3 = st.columns(3)
        col1.metric("Total (ex GST)",  f"${our_total:,.2f}")
        col2.metric("GST",             f"${our_gst:,.2f}")
        col3.metric("Total (inc GST)", f"${our_inc_gst:,.2f}")

        if match:
            st.success(f"✓ MATCH — Our total matches statement's total of ${client_total:,.2f}")
        else:
            diff = abs(our_inc_gst - client_total)
            st.error(f"⚠ DISCREPANCY — Difference of ${diff:,.2f} vs statement's total of ${client_total:,.2f}")

        # ── Breakdown by State ────────────────────────────────────────────
        st.divider()
        st.subheader("Breakdown by State")
        state_summary = []
        for state in sorted(df['State'].unique()):
            s = df[df['State'] == state]
            state_summary.append({
                'State': state,
                'Rows': len(s),
                'Total (AUD)': round(s['Total (AUD)'].sum(), 2),
                'GST (AUD)': round(s['GST (AUD)'].sum(), 2),
                'Total (inc GST)': round(s['Total (AUD)'].sum() + s['GST (AUD)'].sum(), 2)
            })
        st.dataframe(pd.DataFrame(state_summary), use_container_width=True, hide_index=True)

        # ── VIP breakdown ─────────────────────────────────────────────────
        st.divider()
        st.subheader("VIP / Special Service Breakdown")

        vip_df = df[df['VIP'] == True].copy()

        if vip_df.empty:
            if uploaded_csvs:
                st.info("No VIP or special service shipments detected.")
            else:
                st.info("Upload consignment CSV(s) above to detect VIP/special service shipments.")
        else:
            st.caption(
                f"**{len(vip_df)} shipment(s)** identified as VIP/special — "
                "uncheck any that should be treated as general service."
            )

            # Session state for exclusions
            if 'vip_excluded' not in st.session_state:
                st.session_state['vip_excluded'] = set()

            display_cols = ['State', 'Delivery/Adjustment', 'Customer Name & Address',
                            'Service', 'Delivery Special Instructions',
                            'Total (AUD)', 'GST (AUD)']
            display_cols = [c for c in display_cols if c in vip_df.columns]

            checked_rows = []
            for idx, row in vip_df.reset_index().iterrows():
                orig_idx = row['index']
                default_checked = orig_idx not in st.session_state['vip_excluded']
                label = (
                    f"**{row.get('Delivery/Adjustment','')}** | "
                    f"{str(row.get('Customer Name & Address',''))[:50]} | "
                    f"{row.get('State','')} | "
                    f"${row.get('Total (AUD)', 0):.2f}"
                )
                checked = st.checkbox(label, value=default_checked, key=f"vip_chk_{orig_idx}")
                if not checked:
                    st.session_state['vip_excluded'].add(orig_idx)
                else:
                    st.session_state['vip_excluded'].discard(orig_idx)
                if checked:
                    checked_rows.append(row)

            if checked_rows:
                active_vip = pd.DataFrame(checked_rows)[display_cols]
                st.divider()
                st.caption("**Active VIP shipments (included in VIP summary):**")

                # Group by State
                vip_state_summary = []
                for state in sorted(active_vip['State'].unique()):
                    sv = active_vip[active_vip['State'] == state]
                    vip_state_summary.append({
                        'State': state,
                        'VIP Shipments': len(sv),
                        'Total (AUD)': round(sv['Total (AUD)'].astype(float).sum(), 2),
                        'GST (AUD)': round(sv['GST (AUD)'].astype(float).sum(), 2),
                        'Total (inc GST)': round(sv['Total (AUD)'].astype(float).sum() + sv['GST (AUD)'].astype(float).sum(), 2)
                    })
                st.dataframe(pd.DataFrame(vip_state_summary), use_container_width=True, hide_index=True)

                with st.expander("Show full VIP shipment list"):
                    st.dataframe(active_vip, use_container_width=True, hide_index=True)

        # ── Download ──────────────────────────────────────────────────────
        st.divider()
        # Apply exclusions to df before building Excel
        if 'vip_excluded' in st.session_state and st.session_state['vip_excluded']:
            df_for_excel = df.copy()
            vip_indices = df[df['VIP'] == True].index
            for i, orig_idx in enumerate(vip_indices):
                if orig_idx in st.session_state['vip_excluded']:
                    df_for_excel.at[orig_idx, 'VIP'] = False
                    df_for_excel.at[orig_idx, 'VIP Reason'] = ''
        else:
            df_for_excel = df

        buffer = build_excel(df_for_excel, client_total)
        filename = f'{vendor}_split.xlsx'

        st.download_button(
            label="⬇️ Download Excel",
            data=buffer,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# ════════════════════════════════════════════════════════════════════════════
# TAB 2
# ════════════════════════════════════════════════════════════════════════════

with tab2:
    st.write(
        "Compare the **carrier invoice (.TXT)** against the **agreed sell rate card** to catch over/undercharges. "
        "Optionally add the **Machship shipment extract (.CSV)** to also check invoiced sell totals."
    )
    st.caption("🔒 Online build — nothing is stored. All files must be uploaded each session; they're held in memory only.")

    # ── Uploaders ─────────────────────────────────────────────────────────
    c1, c2, c3 = st.columns(3)
    with c1:
        recon_txt = st.file_uploader("① Carrier invoice (.TXT)", type=['txt', 'TXT'], key="tab2_txt")
    with c2:
        recon_csvs = st.file_uploader(
            "② Machship extract (.CSV) — optional",
            type=['csv', 'CSV'], key="tab2_csv", accept_multiple_files=True
        )
    with c3:
        rate_card_file = st.file_uploader("③ Sell rate card (.xlsx) — required", type=['xlsx', 'XLSX'], key="tab2_ratecard")

    # ── Rate card: must be uploaded every session (no storage) ────────────
    rate_card = None
    if rate_card_file is not None:
        try:
            rate_card = parse_rate_card(rate_card_file)
            st.success(
                f"✓ Rate card loaded (this session only) — effective {rate_card.get('effective','?')}, "
                f"metro pallet bands {rate_card.get('metro_band')}."
            )
        except Exception as e:
            st.error(f"⚠ Could not parse rate card: {e}")
    else:
        st.warning("Upload the sell rate card in ③ to enable the checks. It is not stored, so it must be uploaded each session.")

    # ── FSC (hardcoded, not confidential) ─────────────────────────────────
    st.divider()
    st.subheader("⛽ Fuel Surcharge (FSC)")
    st.caption(
        "FSC is hardcoded in the app (the only non-confidential metric). "
        "To change it permanently, edit `CURRENT_FSC_PCT` / `FSC_OVERRIDES` at the top of the file. "
        "You can also override it just for this session below."
    )

    fcol1, fcol2 = st.columns([1, 3])
    with fcol1:
        session_default = st.number_input(
            "Default FSC % (this session)", min_value=0.0, max_value=100.0, step=0.01,
            value=float(CURRENT_FSC_PCT), key="fsc_session_default"
        )
    with fcol2:
        st.caption(
            f"Hardcoded default is **{CURRENT_FSC_PCT}%**. The value here applies to every line "
            "whose date isn't covered by a hardcoded date-range override (shown below). "
            "Changing it here affects this session only — nothing is saved."
        )

    fsc_df = get_fsc_overrides_df()
    default_fsc = float(session_default)

    if not fsc_df.empty:
        st.markdown("**Hardcoded date-range overrides:**")
        show = fsc_df[['date_from', 'date_to', 'fsc_pct', 'note']].rename(
            columns={'date_from': 'From', 'date_to': 'To', 'fsc_pct': 'FSC %', 'note': 'Note'})
        st.dataframe(show, use_container_width=True, hide_index=True)
    else:
        st.caption("No date-range overrides configured — the default FSC applies to all lines.")

    # ── Run reconciliation ────────────────────────────────────────────────
    if recon_txt and rate_card:
        if recon_txt.size > 20 * 1024 * 1024:
            st.error("⚠ TXT file exceeds the 20MB limit.")
            st.stop()

        with st.spinner("Reconciling against rate card..."):
            df_txt, client_total = parse_txt(recon_txt)
            recon_df = reconcile_txt(df_txt, rate_card, fsc_df, default_fsc)

            df_csv_combined = parse_consignment_csvs(recon_csvs) if recon_csvs else pd.DataFrame()
            csv_recon_df = reconcile_csv(df_csv_combined, recon_df) if not df_csv_combined.empty else pd.DataFrame()

        # Helper: implied FSC from invoice for cross-checking
        st.divider()
        with st.expander("🔎 Implied FSC from invoice (LEVY ÷ Freight) — to cross-check the hardcoded FSC"):
            tmp = df_txt.copy()
            tmp['_f'] = tmp['Freight'].apply(clean_number)
            tmp['_l'] = tmp['LEVY'].apply(clean_number)
            tmp = tmp[tmp['_f'] > 0]
            tmp['Implied FSC %'] = (tmp['_l'] / tmp['_f'] * 100).round(2)
            implied = tmp.groupby('Date')['Implied FSC %'].agg(lambda s: ', '.join(map(str, sorted(s.unique())))).reset_index()
            st.dataframe(implied, use_container_width=True, hide_index=True)

        # ── Metrics ────────────────────────────────────────────────────────
        auto = recon_df[recon_df['Check'].str.startswith('Auto')]
        mismatches = recon_df[recon_df['Status'] == '❌ Mismatch']
        st.divider()
        st.subheader("① Invoice (.TXT) vs Rate Card")
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Invoice lines", len(recon_df))
        m2.metric("Auto-checked", len(auto))
        m3.metric("Mismatches", len(mismatches))
        manual_n = (recon_df['Status'] == '🔍 Manual review').sum()
        m4.metric("Manual review", int(manual_n))

        auto_var = auto.dropna(subset=['Total Diff'])
        if not auto_var.empty:
            tot_chg = round(auto_var['Charged Total'].sum(), 2)
            tot_exp = round(auto_var['Expected Total'].sum(), 2)
            var = round(tot_chg - tot_exp, 2)
            v1, v2, v3 = st.columns(3)
            v1.metric("Charged (auto lines)",  f"${tot_chg:,.2f}")
            v2.metric("Expected (auto lines)", f"${tot_exp:,.2f}")
            v3.metric("Variance", f"${var:,.2f}", delta=f"{'Overcharged' if var>0 else 'Undercharged' if var<0 else 'Even'}")
        else:
            tot_chg = tot_exp = var = 0.0

        # ── Highlighted table ────────────────────────────────────────────────
        st.caption("Rows where the invoice differs from the rate card are highlighted red. 🔍 = needs manual review (interstate / regional / tonne).")

        display_cols = ['Status', 'Check', 'Date', 'Delivery/Adjustment', 'Customer Name & Address',
                        'Zone', 'UOM', 'Qty',
                        'Charged Rate', 'Expected Rate', 'Rate Diff',
                        'Charged Freight', 'Expected Freight', 'Freight Diff',
                        'FSC %', 'FSC Source', 'Charged LEVY', 'Expected LEVY', 'LEVY Diff',
                        'Charged Total', 'Expected Total', 'Total Diff']
        display_cols = [c for c in display_cols if c in recon_df.columns]

        def style_recon(df):
            styles = pd.DataFrame('', index=df.index, columns=df.columns)
            for i in range(len(df)):
                status = df.iloc[i]['Status']
                if status == '❌ Mismatch':
                    styles.iloc[i, :] = 'background-color: #fde2e2; color: #8a1f1f'
                elif status == '🔍 Manual review':
                    styles.iloc[i, :] = 'background-color: #fff3cd; color: #7a5b00'
                elif status == '⚠ No FSC for date':
                    styles.iloc[i, :] = 'background-color: #ffe8c2; color: #7c4a00'
                elif status == '✓ OK':
                    styles.iloc[i, :] = 'background-color: #dcf5e3; color: #1f6b3a'
            return styles

        st.dataframe(recon_df[display_cols].style.apply(style_recon, axis=None),
                     use_container_width=True, hide_index=True)

        # ── CSV vs rate card ──────────────────────────────────────────────────
        if not csv_recon_df.empty:
            st.divider()
            st.subheader("② Machship (.CSV) Sell Totals vs Rate Card")
            matched_csv = csv_recon_df[csv_recon_df['Rate Card Expected'].notna()]
            csv_mism = matched_csv[matched_csv['Status'] == '❌ Mismatch']
            cm1, cm2, cm3 = st.columns(3)
            cm1.metric("Invoice shipments in CSV", len(csv_recon_df))
            cm2.metric("Comparable to rate card", len(matched_csv))
            cm3.metric("Mismatches", len(csv_mism))
            st.caption("Only shipments on the invoice (.TXT) are shown. Compares CSV 'Total Sell' to the rate-card expected total; rate-card lines (auto metro) are comparable, interstate/regional show 🔍. Mismatches highlighted red.")

            def style_csv(df):
                styles = pd.DataFrame('', index=df.index, columns=df.columns)
                for i in range(len(df)):
                    s = df.iloc[i]['Status']
                    if s == '❌ Mismatch':
                        styles.iloc[i, :] = 'background-color: #fde2e2; color: #8a1f1f'
                    elif s == '🔍 Manual review':
                        styles.iloc[i, :] = 'background-color: #fff3cd; color: #7a5b00'
                    elif s == '✓ OK':
                        styles.iloc[i, :] = 'background-color: #dcf5e3; color: #1f6b3a'
                return styles

            st.dataframe(csv_recon_df.style.apply(style_csv, axis=None),
                         use_container_width=True, hide_index=True)

        # ── Download Excel ──────────────────────────────────────────────────
        st.divider()

        def build_recon_card_excel(recon_df, csv_recon_df):
            wb = Workbook(); wb.remove(wb.active)
            red  = PatternFill('solid', start_color='C00000')
            grn  = PatternFill('solid', start_color='1F7A1F')
            yel  = PatternFill('solid', start_color='B8860B')
            hdrf = PatternFill('solid', start_color=HEADER_COLOR)

            def write_df(ws, df, title):
                ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
                cols = list(df.columns)
                for c, col in enumerate(cols, 1):
                    cell = ws.cell(row=2, column=c, value=col)
                    cell.font = Font(bold=True, color='FFFFFF'); cell.fill = hdrf
                for r, (_, row) in enumerate(df.iterrows(), 3):
                    status = row.get('Status', '')
                    for c, col in enumerate(cols, 1):
                        val = row[col]
                        if isinstance(val, float): val = round(val, 2)
                        cell = ws.cell(row=r, column=c, value=val)
                        if status == '❌ Mismatch':
                            cell.fill = red; cell.font = Font(color='FFFFFF')
                        elif status == '🔍 Manual review':
                            cell.fill = yel; cell.font = Font(color='FFFFFF')
                        elif status == '✓ OK':
                            cell.fill = grn; cell.font = Font(color='FFFFFF')

            write_df(wb.create_sheet("TXT vs Rate Card"), recon_df, "Invoice (.TXT) vs Rate Card")
            if not csv_recon_df.empty:
                write_df(wb.create_sheet("CSV vs Rate Card"), csv_recon_df, "Machship (.CSV) vs Rate Card")
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            return buf

        excel_buf = build_recon_card_excel(recon_df, csv_recon_df)
        st.download_button(
            "⬇️ Download Reconciliation Report (.xlsx)",
            data=excel_buf,
            file_name="rate_card_reconciliation.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# ════════════════════════════════════════════════════════════════════════════
# TAB 3 — Screenshot to Excel (free OCR via pytesseract)
# ════════════════════════════════════════════════════════════════════════════

with tab3:
    st.write("Upload one or more screenshots of your charge table and download the extracted data as Excel.")

    screenshots = st.file_uploader(
        "Upload screenshots (.png, .jpg)",
        type=["png", "jpg", "jpeg"],
        accept_multiple_files=True,
        key="tab3_screenshots"
    )

    COLUMNS = [
        "Charge", "Description", "Bran", "Depa", "Cost",
        "OS Cost Amt", "Estimated Cost", "Local Cost Amt", "Creditor",
        "Cost Recognition", "Posted", "Apt", "Sell", "OS Sell Amt",
        "Estimated Revenue", "Local Sell Amt", "Debtor",
        "Sell Recognition", "Posted (Sell)"
    ]

    COL_X = [109, 256, 799, 883, 978, 1072, 1289, 1521, 1759,
              1924, 2176, 2293, 2417, 2525, 2728, 2967, 3152, 3352, 3596]

    def assign_col(left):
        dists = [abs(left - cx) for cx in COL_X]
        return COLUMNS[dists.index(min(dists))]

    def clean_charge(val):
        if not isinstance(val, str): return val
        m = re.search(r'\b(W[A-Z]{3})\b', val)
        return m.group(1) if m else val.strip()

    def clean_numeric_ocr(val):
        if not isinstance(val, str): return None
        val = val.replace(",", ".").strip()
        try: return float(val)
        except: return None

    def extract_table_ocr(image_bytes):
        from PIL import Image, ImageEnhance
        import pytesseract

        img = Image.open(io.BytesIO(image_bytes))
        w, h = img.size
        img = img.resize((w * 2, h * 2), Image.LANCZOS)
        img = ImageEnhance.Contrast(img).enhance(1.5)

        df_ocr = pytesseract.image_to_data(img, output_type=pytesseract.Output.DATAFRAME, config="--psm 6")
        df_ocr = df_ocr[df_ocr["conf"] > 0].copy()
        df_ocr["text"] = df_ocr["text"].astype(str).str.strip()
        df_ocr = df_ocr[df_ocr["text"].str.len() > 0].reset_index(drop=True)

        df_ocr = df_ocr.sort_values("top")
        row_labels, current_label, last_top = [], 0, -100
        for _, r in df_ocr.iterrows():
            if r["top"] - last_top > 30:
                current_label += 1
                last_top = r["top"]
            row_labels.append(current_label)
        df_ocr["row_id"] = row_labels

        rows_out = []
        for rid in sorted(df_ocr["row_id"].unique()):
            if rid == 1: continue
            row_words = df_ocr[df_ocr["row_id"] == rid].sort_values("left")
            row_dict = {}
            for _, w in row_words.iterrows():
                col = assign_col(w["left"])
                row_dict[col] = (row_dict.get(col, "") + " " + w["text"]).strip()
            if row_dict:
                rows_out.append(row_dict)

        return pd.DataFrame(rows_out, columns=COLUMNS)

    if screenshots:
        if st.button("🔍 Extract Table Data", key="tab3_extract"):
            all_rows = []
            progress = st.progress(0)
            status   = st.empty()

            for i, img_file in enumerate(screenshots):
                status.text(f"Processing {img_file.name} ({i+1}/{len(screenshots)})...")
                try:
                    img_bytes = img_file.read()
                    df_img    = extract_table_ocr(img_bytes)
                    df_img["_source"] = img_file.name
                    all_rows.append(df_img)
                except Exception as e:
                    st.warning(f"⚠ Could not extract from {img_file.name}: {e}")
                progress.progress((i + 1) / len(screenshots))

            status.empty()
            progress.empty()

            if all_rows:
                df_shots = pd.concat(all_rows, ignore_index=True)
                df_shots["Charge"] = df_shots["Charge"].apply(clean_charge)
                for col in ["OS Cost Amt", "Estimated Cost", "Local Cost Amt",
                            "OS Sell Amt", "Estimated Revenue", "Local Sell Amt"]:
                    df_shots[col] = df_shots[col].apply(clean_numeric_ocr)
                    df_shots[col] = pd.to_numeric(df_shots[col], errors="coerce")
                st.session_state["screenshot_df"] = df_shots
                st.success(f"✓ Extracted {len(df_shots)} rows from {len(screenshots)} screenshot(s)")

    if "screenshot_df" in st.session_state:
        df_shots     = st.session_state["screenshot_df"]
        display_cols = [c for c in COLUMNS if c in df_shots.columns]

        st.divider()
        st.subheader("Extracted Data")
        st.dataframe(df_shots[display_cols], use_container_width=True, hide_index=True)

        st.divider()
        st.subheader("OS Cost Amt by Creditor")
        cost_summary = (
            df_shots.groupby("Creditor")["OS Cost Amt"].sum().reset_index()
            .rename(columns={"OS Cost Amt": "Total OS Cost Amt"})
            .sort_values("Total OS Cost Amt", ascending=False)
        )
        cost_summary["Total OS Cost Amt"] = cost_summary["Total OS Cost Amt"].round(2)
        st.dataframe(cost_summary, use_container_width=True, hide_index=True)

        st.divider()
        st.subheader("Estimated Revenue by Debtor")
        rev_summary = (
            df_shots.groupby("Debtor")["Estimated Revenue"].sum().reset_index()
            .rename(columns={"Estimated Revenue": "Total Estimated Revenue"})
            .sort_values("Total Estimated Revenue", ascending=False)
        )
        rev_summary["Total Estimated Revenue"] = rev_summary["Total Estimated Revenue"].round(2)
        st.dataframe(rev_summary, use_container_width=True, hide_index=True)

        st.divider()

        def build_screenshot_excel(df, cost_sum, rev_sum):
            wb = Workbook()
            wb.remove(wb.active)

            def write_sheet(ws, df_in, title):
                ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
                for c, col in enumerate(df_in.columns, 1):
                    cell = ws.cell(row=2, column=c, value=col)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", start_color="1F4E79")
                for r, (_, row) in enumerate(df_in.iterrows(), 3):
                    for c, val in enumerate(row, 1):
                        ws.cell(row=r, column=c, value=val)

            write_sheet(wb.create_sheet("Extracted Data"),    df[display_cols], "Extracted Charge Data")
            write_sheet(wb.create_sheet("Cost by Creditor"),  cost_sum,         "OS Cost Amt by Creditor")
            write_sheet(wb.create_sheet("Revenue by Debtor"), rev_sum,          "Estimated Revenue by Debtor")

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf

        excel_buf = build_screenshot_excel(df_shots, cost_summary, rev_summary)
        st.download_button(
            label="⬇️ Download Excel",
            data=excel_buf,
            file_name="screenshot_charges.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# ════════════════════════════════════════════════════════════════════════════
# TAB 4 — Master Consolidation
# ════════════════════════════════════════════════════════════════════════════

with tab4:
    st.write(
        "Upload **multiple client statement `.TXT` files** (include a previously "
        "downloaded `master_file.txt` plus any new invoices) and **multiple Machship "
        "`.CSV` extracts**. Everything is combined into one workbook — split by state, "
        "consolidated per 1800 number, and compared against what Machship says we invoiced."
    )
    st.caption("🔒 Online build — nothing is stored. Download the master_file.txt and re-upload it "
               "next time to keep growing it.")

    mc1, mc2, mc3 = st.columns(3)
    with mc1:
        master_txts = st.file_uploader(
            "Client statements (.TXT) — master_file.txt + new invoices",
            type=['txt', 'TXT'], key="tab4_txts", accept_multiple_files=True)
    with mc2:
        master_csvs = st.file_uploader(
            "Machship extracts (.CSV) — optional, enables the comparison",
            type=['csv', 'CSV'], key="tab4_csvs", accept_multiple_files=True)
    with mc3:
        master_ratecard = st.file_uploader(
            "Sell rate card (.xlsx) — optional, source of truth",
            type=['xlsx', 'XLSX'], key="tab4_ratecard")

    if master_txts:
        oversize = [f.name for f in master_txts if f.size > 20 * 1024 * 1024]
        if oversize:
            st.error(f"⚠ These files exceed the 20MB limit: {', '.join(oversize)}")
            st.stop()

        with st.spinner("Combining statements..."):
            df_all, client_total, info = parse_multiple_txt(master_txts)

        if df_all.empty:
            st.error("No valid data found in the uploaded .TXT files.")
            st.stop()

        st.divider()
        st.subheader("Uploaded statements")
        st.dataframe(pd.DataFrame(info), use_container_width=True, hide_index=True)

        # Exact-duplicate detection (protects against re-uploading a file already in the master)
        dup_mask = df_all.duplicated(subset=TXT_COLUMNS, keep=False)
        n_dup = int(dup_mask.sum())
        if n_dup:
            st.warning(f"⚠ {n_dup} line(s) are exact duplicates across your uploads — this happens "
                       "if a statement already inside the master was uploaded again.")
            if st.checkbox("Remove exact duplicate lines before consolidating", value=False,
                           key="tab4_dedupe"):
                before = len(df_all)
                df_all = df_all.drop_duplicates(subset=TXT_COLUMNS, keep='first').reset_index(drop=True)
                client_total = round(df_all['Total (AUD)'].sum(), 2)
                st.info(f"Removed {before - len(df_all)} duplicate line(s); {len(df_all)} remain.")

        # ── Machship extracts + column mapping ────────────────────────────
        ref_col = sell_col = None
        df_csv = pd.DataFrame()
        if master_csvs:
            df_csv = parse_consignment_csvs(master_csvs)
            if not df_csv.empty:
                st.divider()
                st.subheader("Machship column mapping")
                auto_ref, auto_sell = detect_ref_column(df_csv), detect_sell_column(df_csv)
                opts = list(df_csv.columns)
                colr, cols_ = st.columns(2)
                with colr:
                    ref_col = st.selectbox(
                        "Reference column (contains the 1800 number)", options=opts,
                        index=(opts.index(auto_ref) if auto_ref in opts else 0), key="tab4_refcol")
                with cols_:
                    sell_col = st.selectbox(
                        "Sell amount column — ex GST (what we invoice)", options=opts,
                        index=(opts.index(auto_sell) if auto_sell in opts else 0), key="tab4_sellcol")
                st.caption("Pick the **ex-GST** sell/charge column so it lines up with the client's "
                           "Total (AUD), which is also ex GST.")

        # ── Rate card (optional) + hardcoded sell fuel surcharge ──────────
        rate_card = None
        if master_ratecard is not None:
            try:
                master_ratecard.seek(0)
            except Exception:
                pass
            try:
                rate_card = parse_rate_card(master_ratecard)
                st.divider()
                st.success(f"✓ Rate card loaded (this session only) — effective "
                           f"{rate_card.get('effective', '?')}, metro pallet bands "
                           f"{rate_card.get('metro_band')}.")
            except Exception as e:
                st.error(f"⚠ Could not parse rate card: {e}")
                rate_card = None

        with st.expander("⛽ Sell Fuel Surcharge (hardcoded — no upload needed)", expanded=False):
            st.caption("Applied to the rate-card 'should have billed' figure, by each line's date. "
                       "Edit `SELL_FSC_EFFECTIVE` / `SELL_FSC_BASE_PCT` at the top of the file to change it.")
            st.dataframe(get_sell_fsc_df(), use_container_width=True, hide_index=True)

        # ── Build consolidation + comparisons ─────────────────────────────
        consol_df = consolidate_by_ref(df_all)
        mship_df  = (machship_by_ref(df_csv, ref_col, sell_col)
                     if not df_csv.empty else
                     pd.DataFrame(columns=['Ref', 'Machship Sell (ex GST)', 'Machship Rows']))
        rc_df     = (compute_rate_card_expected(df_csv, rate_card, ref_col)
                     if (rate_card and not df_csv.empty and ref_col) else None)
        disc_df   = build_discrepancy(consol_df, mship_df, rc_df)
        state_cmp = build_state_comparison(consol_df, mship_df, rc_df)

        # ── Metrics ───────────────────────────────────────────────────────
        st.divider()
        k1, k2, k3, k4 = st.columns(4)
        k1.metric("Statements", len(info))
        k2.metric("Invoice lines", len(df_all))
        k3.metric("Unique 1800 refs", len(consol_df))
        k4.metric("Client total (ex GST)", f"${df_all['Total (AUD)'].sum():,.2f}")

        # ── Per-state comparison ──────────────────────────────────────────
        st.divider()
        st.subheader("Per state — what we thought we make vs their figure (ex GST)")
        st.caption("'We Thought' = Machship sell (ex GST). 'Their Figure' = consolidated client Total (AUD), ex GST.")
        st.dataframe(state_cmp if not state_cmp.empty else pd.DataFrame({'info': ['Upload Machship CSV(s) to compare']}),
                     use_container_width=True, hide_index=True)

        # ── Discrepancies ─────────────────────────────────────────────────
        if not disc_df.empty and not mship_df.empty:
            st.divider()
            st.subheader("Discrepancies per 1800 number (ex GST)")
            n_mis = int((disc_df['Status'] == '❌ Discrepancy').sum())
            n_noms = int((disc_df['Status'] == '⚠ No Machship data').sum())
            n_nocl = int((disc_df['Status'] == '⚠ No client line').sum())
            st.caption(f"{n_mis} mismatch(es) · {n_noms} client-only (no Machship) · {n_nocl} Machship-only (no client line).")
            st.dataframe(disc_df, use_container_width=True, hide_index=True)
        elif master_csvs:
            st.info("Choose a valid Machship reference and ex-GST sell column above to see the comparison.")
        else:
            st.info("Add Machship CSV extract(s) to compare their figures against what we invoiced.")

        # ── Consolidated per 1800 ─────────────────────────────────────────
        st.divider()
        st.subheader("Consolidated cost per 1800 number")
        st.caption("Charges and credits (negative costs) net off within each 1800 number to give the final cost paid.")
        st.dataframe(consol_df, use_container_width=True, hide_index=True)

        # ── Downloads ─────────────────────────────────────────────────────
        st.divider()
        d1, d2 = st.columns(2)
        with d1:
            st.download_button(
                "⬇️ Download master_file.txt (re-upload next time)",
                data=build_master_txt(df_all),
                file_name="master_file.txt", mime="text/plain")
            st.caption("Combined client data + trailing total, in the original .TXT format.")
        with d2:
            xbuf = build_master_excel(df_all, client_total, consol_df, mship_df, disc_df, state_cmp)
            st.download_button(
                "⬇️ Download master workbook (.xlsx)",
                data=xbuf, file_name="master_consolidated.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.caption("Summary, raw_data, per-state, consolidated_by_1800, discrepancies, state_vs_machship.")

# ════════════════════════════════════════════════════════════════════════════
# TAB 5 — TGE Mapper
# ════════════════════════════════════════════════════════════════════════════

with tab5:
    st.write(
        "Upload one or more **TGE invoice PDFs** and the **Machship consignment export (.CSV)**. "
        "Each invoice shipment is mapped to the consignment export so the **Machship #** and "
        "**Reference 2 (the W-number)** get pulled onto every invoice line."
    )
    st.caption(
        "Matching key: invoice **Customer Reference 1** → consignment **Reference 1** "
        "(with a **Machship #** fallback, since the reference entered is often the Machship number itself). "
        "Consignments that are on the export but not on the invoice are ignored."
    )

    mcol1, mcol2 = st.columns(2)
    with mcol1:
        tge_invoices = st.file_uploader(
            "① Invoice PDFs (e.g. 1269984.pdf, 1272416.pdf)",
            type=['pdf', 'PDF'], key="tab5_invoices", accept_multiple_files=True)
    with mcol2:
        tge_csvs = st.file_uploader(
            "② Machship consignment export (.CSV)",
            type=['csv', 'CSV'], key="tab5_csvs", accept_multiple_files=True)

    if tge_invoices and tge_csvs:
        # Make sure pdfplumber is available before doing anything.
        try:
            import pdfplumber  # noqa: F401
            _has_pdf = True
        except Exception:
            _has_pdf = False

        if not _has_pdf:
            st.error("This tab needs the **pdfplumber** package to read invoice PDFs. "
                     "Install it with `pip install pdfplumber` and reload.")
        else:
            oversize = [f.name for f in tge_invoices if f.size > 20 * 1024 * 1024]
            if oversize:
                st.error(f"⚠ These PDFs exceed the 20MB limit: {', '.join(oversize)}")
                st.stop()

            with st.spinner("Reading invoices and matching against the consignment export..."):
                invoice_rows, parse_errors = [], []
                for f in tge_invoices:
                    try:
                        rows, _inv = parse_invoice_pdf(f, getattr(f, 'name', 'invoice.pdf'))
                        invoice_rows += rows
                    except Exception as e:
                        parse_errors.append(f"{getattr(f, 'name', 'a PDF')}: {e}")
                df_csv = parse_consignment_csvs(tge_csvs)
                df_map = build_tge_mapping(invoice_rows, df_csv) if invoice_rows else pd.DataFrame()

            for err in parse_errors:
                st.warning(f"⚠ Could not read {err}")

            if df_map.empty:
                st.error("No invoice line items could be read from the uploaded PDF(s). "
                         "Make sure these are the TGE 'Intermodal & Specialised' invoices.")
            else:
                total     = len(df_map)
                matched   = int(df_map['_matched'].sum())
                unmatched = total - matched

                # ── First line: were all invoice lines found? ──────────────
                if unmatched == 0:
                    st.success(f"✅ All {total} invoice line(s) were matched to the consignment export.")
                else:
                    st.warning(
                        f"⚠️ {matched} of {total} invoice line(s) matched — "
                        f"**{unmatched} could not be matched** and are highlighted orange below "
                        "so you can double-check where these shipments are."
                    )

                m1, m2, m3 = st.columns(3)
                m1.metric("Invoice lines", total)
                m2.metric("Matched", matched)
                m3.metric("Unmatched", unmatched)

                order   = tge_display_order(df_map)
                show_df = df_map[order].copy()
                unmatched_mask = (~df_map['_matched']).tolist()

                # ── Mapped table (unmatched rows highlighted orange) ───────
                st.divider()
                st.subheader("Invoice lines → consignment export")
                st.caption("Added columns **Machship #** and **Reference 2** come from the consignment export. "
                           "Orange rows couldn't be matched and have blank added cells.")

                def _highlight_unmatched(_):
                    styles = pd.DataFrame('', index=show_df.index, columns=show_df.columns)
                    for i in range(len(show_df)):
                        if unmatched_mask[i]:
                            styles.iloc[i, :] = 'background-color: #ffe3c2; color: #7c3a00'
                    return styles

                st.dataframe(show_df.style.apply(_highlight_unmatched, axis=None),
                             use_container_width=True, hide_index=True)

                if unmatched:
                    with st.expander(f"Show only the {unmatched} unmatched line(s)"):
                        st.dataframe(show_df[unmatched_mask], use_container_width=True, hide_index=True)

                # ── Download Excel ─────────────────────────────────────────
                st.divider()
                xbuf = build_tge_mapper_excel(df_map, order)
                st.download_button(
                    "⬇️ Download mapped Excel",
                    data=xbuf, file_name="tge_mapped.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                st.caption("One sheet: every invoice line with Machship # and Reference 2 added; "
                           "unmatched rows filled orange.")
    else:
        st.info("Upload at least one invoice PDF and one consignment CSV to run the mapping.")
